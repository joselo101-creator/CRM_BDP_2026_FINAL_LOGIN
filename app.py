try:
    import cgi
except Exception:
    cgi = None
import csv
import hashlib
import hmac
import io
import json
import os
import shutil
import smtplib
import sqlite3
import threading
import unicodedata
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qs, urlparse

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except Exception:
    colors = None
    letter = None
    getSampleStyleSheet = None
    Image = None
    Paragraph = None
    SimpleDocTemplate = None
    Spacer = None
    Table = None
    TableStyle = None


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", str(APP_DIR))).resolve()
DB_PATH = DATA_DIR / "bdp_crm.sqlite"
BACKUP_DIR = DATA_DIR / "backups"
QUOTE_DIR = DATA_DIR / "cotizaciones_pdf_actuales"
QUOTE_WON_DIR = DATA_DIR / "cotizaciones_logradas"
SOURCE_XLSX = Path(r"C:\Users\Equipo\Documents\ANALISIS Y REPORTES 2026 BDP\control_visitas_red_distribuidores_nayarit.xlsx")


def bootstrap_persistent_data():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    seed_files = [
        (APP_DIR / "bdp_crm.sqlite", DB_PATH),
    ]
    for src, dst in seed_files:
        if src.exists() and not dst.exists():
            shutil.copy2(src, dst)
    for folder_name in ("backups", "cotizaciones_pdf_actuales", "cotizaciones_logradas"):
        src_dir = APP_DIR / folder_name
        dst_dir = DATA_DIR / folder_name
        dst_dir.mkdir(parents=True, exist_ok=True)
        if src_dir.exists():
            for src in src_dir.iterdir():
                dst = dst_dir / src.name
                if src.is_file() and not dst.exists():
                    shutil.copy2(src, dst)

REGIONES = {
    "Costa Sur": {
        "color": "#1f8f73",
        "descripcion": "Region de alto desarrollo turistico y hotelero internacional.",
        "municipios": ["Bahía de Banderas", "Compostela"],
    },
    "Costa Centro y Valles del Norte": {
        "color": "#2d76a3",
        "descripcion": "Zonas costeras, agricolas y pesqueras del centro-norte.",
        "municipios": ["San Blas", "Santiago Ixcuintla", "Tuxpan", "Ruiz"],
    },
    "Costa Norte y Frontera": {
        "color": "#d17a21",
        "descripcion": "Llanura costera del norte, pesca de camaron y agricultura.",
        "municipios": ["Tecuala", "Acaponeta", "Huajicori", "Rosamorada"],
    },
    "Centro y Sierra Centro": {
        "color": "#7a5ac8",
        "descripcion": "Valle central, capital y area conurbada.",
        "municipios": ["Tepic", "Xalisco", "Santa María del Oro", "San Pedro Lagunillas"],
    },
    "Sierra y Sur": {
        "color": "#8b6f3d",
        "descripcion": "Sierra Madre Occidental, minas, volcanes y corredor sur.",
        "municipios": ["Del Nayar", "La Yesca", "Jala", "Ahuacatlán", "Ixtlán del Río", "Amatlán de Cañas"],
    },
}

MUNICIPIO_REGION = {municipio: region for region, data in REGIONES.items() for municipio in data["municipios"]}

NAYARIT_MUNICIPIOS = {
    "Acaponeta": {"cabecera": "Acaponeta", "zona": "Costa Norte y Frontera"},
    "Ahuacatlán": {"cabecera": "Ahuacatlán", "zona": "Sierra y Sur"},
    "Amatlán de Cañas": {"cabecera": "Amatlán de Cañas", "zona": "Sierra y Sur"},
    "Bahía de Banderas": {"cabecera": "Valle de Banderas", "zona": "Costa Sur"},
    "Compostela": {"cabecera": "Compostela", "zona": "Costa Sur"},
    "Del Nayar": {"cabecera": "Jesús María", "zona": "Sierra y Sur"},
    "Huajicori": {"cabecera": "Huajicori", "zona": "Costa Norte y Frontera"},
    "Ixtlán del Río": {"cabecera": "Ixtlán del Río", "zona": "Sierra y Sur"},
    "Jala": {"cabecera": "Jala", "zona": "Sierra y Sur"},
    "La Yesca": {"cabecera": "La Yesca", "zona": "Sierra y Sur"},
    "Rosamorada": {"cabecera": "Rosamorada", "zona": "Costa Norte y Frontera"},
    "Ruiz": {"cabecera": "Ruiz", "zona": "Costa Centro y Valles del Norte"},
    "San Blas": {"cabecera": "San Blas", "zona": "Costa Centro y Valles del Norte"},
    "San Pedro Lagunillas": {"cabecera": "San Pedro Lagunillas", "zona": "Centro y Sierra Centro"},
    "Santa María del Oro": {"cabecera": "Santa María del Oro", "zona": "Centro y Sierra Centro"},
    "Santiago Ixcuintla": {"cabecera": "Santiago Ixcuintla", "zona": "Costa Centro y Valles del Norte"},
    "Tecuala": {"cabecera": "Tecuala", "zona": "Costa Norte y Frontera"},
    "Tepic": {"cabecera": "Tepic", "zona": "Centro y Sierra Centro"},
    "Tuxpan": {"cabecera": "Tuxpan", "zona": "Costa Centro y Valles del Norte"},
    "Xalisco": {"cabecera": "Xalisco", "zona": "Centro y Sierra Centro"},
}

SAT_CATALOGS = {
    "regimenes": {
        "Fisica": [
            ["605", "Sueldos y Salarios e Ingresos Asimilados a Salarios"],
            ["606", "Arrendamiento"],
            ["607", "Regimen de Enajenacion o Adquisicion de Bienes"],
            ["608", "Demas ingresos"],
            ["611", "Ingresos por Dividendos"],
            ["612", "Personas Fisicas con Actividades Empresariales y Profesionales"],
            ["614", "Ingresos por intereses"],
            ["615", "Regimen de los ingresos por obtencion de premios"],
            ["616", "Sin obligaciones fiscales"],
            ["621", "Incorporacion Fiscal"],
            ["625", "Regimen de las Actividades Empresariales con ingresos a traves de Plataformas Tecnologicas"],
            ["626", "Regimen Simplificado de Confianza"],
        ],
        "Moral": [
            ["601", "General de Ley Personas Morales"],
            ["603", "Personas Morales con Fines no Lucrativos"],
            ["610", "Residentes en el Extranjero sin Establecimiento Permanente en Mexico"],
            ["620", "Sociedades Cooperativas de Produccion que optan por diferir sus ingresos"],
            ["622", "Actividades Agricolas, Ganaderas, Silvicolas y Pesqueras"],
            ["623", "Opcional para Grupos de Sociedades"],
            ["624", "Coordinados"],
            ["626", "Regimen Simplificado de Confianza"],
        ],
    },
    "usos_cfdi": {
        "Fisica": [
            ["G01", "Adquisicion de mercancias"],
            ["G03", "Gastos en general"],
            ["I01", "Construcciones"],
            ["I02", "Mobiliario y equipo de oficina por inversiones"],
            ["I03", "Equipo de transporte"],
            ["I04", "Equipo de computo y accesorios"],
            ["I05", "Dados, troqueles, moldes, matrices y herramental"],
            ["I08", "Otra maquinaria y equipo"],
            ["D01", "Honorarios medicos, dentales y gastos hospitalarios"],
            ["D02", "Gastos medicos por incapacidad o discapacidad"],
            ["D03", "Gastos funerales"],
            ["D04", "Donativos"],
            ["D05", "Intereses reales efectivamente pagados por creditos hipotecarios"],
            ["D06", "Aportaciones voluntarias al SAR"],
            ["D07", "Primas por seguros de gastos medicos"],
            ["D08", "Gastos de transportacion escolar obligatoria"],
            ["D09", "Depositos en cuentas para el ahorro"],
            ["D10", "Pagos por servicios educativos"],
            ["S01", "Sin efectos fiscales"],
            ["CP01", "Pagos"],
            ["CN01", "Nomina"],
        ],
        "Moral": [
            ["G01", "Adquisicion de mercancias"],
            ["G03", "Gastos en general"],
            ["I01", "Construcciones"],
            ["I02", "Mobiliario y equipo de oficina por inversiones"],
            ["I03", "Equipo de transporte"],
            ["I04", "Equipo de computo y accesorios"],
            ["I05", "Dados, troqueles, moldes, matrices y herramental"],
            ["I08", "Otra maquinaria y equipo"],
            ["S01", "Sin efectos fiscales"],
            ["CP01", "Pagos"],
        ],
    },
    "metodos_pago": [["PUE", "Pago en una sola exhibicion"], ["PPD", "Pago en parcialidades o diferido"]],
    "formas_pago": [
        ["01", "Efectivo"], ["02", "Cheque nominativo"], ["03", "Transferencia electronica de fondos"],
        ["04", "Tarjeta de credito"], ["28", "Tarjeta de debito"], ["99", "Por definir"]
    ],
}

PRODUCTOS = {
    "Block 10 × 14 × 28 cm": 5.50,
    "Block 12 × 20 × 40 cm": None,
    "Block 14 × 20 × 40 cm": 14.56,
    "Bovedilla prefabricada": None,
    "Adoquín tipo romano": None,
}

TARIFAS = {
    "Costa Sur": 28.40,
    "Costa Centro y Valles del Norte": 28.40,
    "Costa Norte y Frontera": 28.40,
    "Centro y Sierra Centro": 26.00,
    "Sierra y Sur": 28.40,
}


DEFAULT_SETTINGS = {
    "user_name": "José Luis",
    "user_role": "Administrador",
    "company_name": "Bazar de Prefabricados",
    "annual_goal": "25000000",
    "monthly_goal_default": "2000000",
    "profile_photo": "",
    "quote_admin_email": "bdprefabricados@hotmail.com",
    "smtp_host": "",
    "smtp_port": "587",
    "smtp_user": "",
    "smtp_password": "",
    "smtp_from_email": "",
    "smtp_from_name": "BAZAR DE PREFABRICADOS",
    "material_cemento_ton": "4300",
    "material_jal_m3": "1600",
    "mo_block_10_14_28": "0",
    "mo_block_14_20_40": "0",
}

MESES_ES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

def ensure_business_tables(db):
    db.executescript("""
        CREATE TABLE IF NOT EXISTS app_settings (
          key TEXT PRIMARY KEY,
          value TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS monthly_sales (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          year INTEGER NOT NULL,
          month INTEGER NOT NULL,
          amount REAL DEFAULT 0,
          goal REAL DEFAULT 0,
          updated_at TEXT DEFAULT '',
          UNIQUE(year, month)
        );
        CREATE TABLE IF NOT EXISTS product_prices (
          producto TEXT PRIMARY KEY,
          precio REAL,
          updated_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS products (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          nombre TEXT UNIQUE NOT NULL,
          medidas TEXT DEFAULT '',
          resistencia TEXT DEFAULT '',
          peso_aprox TEXT DEFAULT '',
          precio REAL,
          activo INTEGER DEFAULT 1,
          created_at TEXT DEFAULT '',
          updated_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS freight_rates (
          zona TEXT PRIMARY KEY,
          tarifa REAL,
          updated_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS users (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          username TEXT UNIQUE NOT NULL,
          password_hash TEXT NOT NULL,
          full_name TEXT DEFAULT '',
          role TEXT DEFAULT 'Vendedor',
          active INTEGER DEFAULT 1,
          created_at TEXT DEFAULT '',
          updated_at TEXT DEFAULT ''
        );
    """)
    for k, v in DEFAULT_SETTINGS.items():
        db.execute("INSERT OR IGNORE INTO app_settings(key,value) VALUES(?,?)", (k, str(v)))
    year = date.today().year
    default_goal = float(DEFAULT_SETTINGS["monthly_goal_default"])
    for m in range(1,13):
        db.execute("INSERT OR IGNORE INTO monthly_sales(year, month, amount, goal, updated_at) VALUES(?,?,?,?,?)", (year, m, 0, default_goal, now_iso()))
    for prod, precio in PRODUCTOS.items():
        db.execute(
            "INSERT OR IGNORE INTO products(nombre,precio,activo,created_at,updated_at) VALUES(?,?,1,?,?)",
            (prod, precio, now_iso(), now_iso()),
        )
        if precio is not None:
            db.execute("INSERT OR IGNORE INTO product_prices(producto, precio, updated_at) VALUES(?,?,?)", (prod, precio, now_iso()))
    for zona, tarifa in TARIFAS.items():
        db.execute("INSERT OR IGNORE INTO freight_rates(zona, tarifa, updated_at) VALUES(?,?,?)", (zona, tarifa, now_iso()))
    admin = db.execute("SELECT id FROM users WHERE username=?", (auth_user(),)).fetchone()
    if not admin:
        db.execute(
            """
            INSERT INTO users (username, password_hash, full_name, role, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, 1, ?, ?)
            """,
            (auth_user(), hash_password(auth_password()), "Administrador BDP", "Administrador", now_iso(), now_iso()),
        )


def get_settings_payload(db, year=None):
    year = int(year or date.today().year)
    settings = {r["key"]: r["value"] for r in db.execute("SELECT key,value FROM app_settings").fetchall()}
    sales = rows_to_dicts(db.execute("SELECT year, month, amount, goal FROM monthly_sales WHERE year=? ORDER BY month", (year,)).fetchall())
    product_rows = db.execute("SELECT producto, precio FROM product_prices ORDER BY producto").fetchall()
    product_details = rows_to_dicts(db.execute("SELECT * FROM products WHERE activo=1 ORDER BY nombre COLLATE NOCASE").fetchall())
    freight_rows = db.execute("SELECT zona, tarifa FROM freight_rates ORDER BY zona").fetchall()
    products = dict(PRODUCTOS)
    for row in product_details:
        products[row["nombre"]] = row["precio"]
    for r in product_rows:
        products[r["producto"]] = r["precio"]
    for row in product_details:
        row["precio"] = products.get(row["nombre"])
    tarifas = dict(TARIFAS)
    for r in freight_rows:
        tarifas[r["zona"]] = r["tarifa"]
    return {"settings": settings, "monthlySales": sales, "productos": products, "productDetails": product_details, "tarifas": tarifas, "year": year, "months": MESES_ES}



def parse_upload_file(body, content_type):
    """Small multipart/form-data parser used when cgi.FieldStorage is unavailable."""
    if "boundary=" not in content_type:
        return None, "Carga manual"
    boundary = content_type.split("boundary=", 1)[1].strip().strip('"')
    marker = ("--" + boundary).encode("utf-8")
    for part in body.split(marker):
        if b"Content-Disposition" not in part:
            continue
        header, sep, payload = part.partition(b"\r\n\r\n")
        if not sep:
            continue
        header_text = header.decode("utf-8", errors="ignore")
        if 'name="file"' not in header_text:
            continue
        filename = "Carga manual"
        if 'filename="' in header_text:
            filename = header_text.split('filename="', 1)[1].split('"', 1)[0] or filename
        payload = payload.rstrip(b"\r\n")
        if payload.endswith(b"--"):
            payload = payload[:-2].rstrip(b"\r\n")
        return payload, filename
    return None, "Carga manual"

def conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c


def now_iso():
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def normalize_date(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).split(" ")[0]


def zone_for(municipio):
    return NAYARIT_MUNICIPIOS.get(municipio or "", {}).get("zona", "")


def sync_regions(db):
    for municipio, data in NAYARIT_MUNICIPIOS.items():
        db.execute(
            "UPDATE clients SET zona=?, cabecera=? WHERE municipio=?",
            (data["zona"], data["cabecera"], municipio),
        )


def migrate_existing_quote_pdfs(db):
    for folder, status, prefix in ((QUOTE_DIR, "Elaborada", "/cotizaciones_pdf/"), (QUOTE_WON_DIR, "Lograda", "/cotizaciones_logradas/")):
        for pdf in folder.glob("*.pdf"):
            pdf_url = prefix + pdf.name
            exists = db.execute("SELECT id FROM quotes WHERE pdf LIKE ?", (f"%/{pdf.name}",)).fetchone()
            if exists:
                continue
            parts = pdf.stem.split("_")
            client_id = parts[1] if len(parts) > 2 and parts[1].isdigit() else ""
            client = db.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone() if client_id else None
            raw_date = parts[2] if len(parts) > 3 else ""
            fecha = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:8]}" if len(raw_date) == 8 else date.today().isoformat()
            db.execute(
                """
                INSERT INTO quotes
                (client_id, fecha, cliente, producto, status, pdf, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?)
                """,
                (
                    int(client_id) if client_id else 0,
                    fecha,
                    client["nombre_comercial"] if client else pdf.stem,
                    "",
                    status,
                    pdf_url,
                    now_iso(),
                    now_iso(),
                ),
            )


def init_db():
    BACKUP_DIR.mkdir(exist_ok=True)
    QUOTE_DIR.mkdir(exist_ok=True)
    QUOTE_WON_DIR.mkdir(exist_ok=True)
    with conn() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS clients (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              nombre_comercial TEXT NOT NULL,
              razon_social TEXT DEFAULT '',
              tipo_persona_fiscal TEXT DEFAULT '',
              rfc TEXT DEFAULT '',
              regimen_fiscal TEXT DEFAULT '',
              uso_cfdi TEXT DEFAULT '',
              cp_fiscal TEXT DEFAULT '',
              domicilio_fiscal TEXT DEFAULT '',
              correo_facturacion TEXT DEFAULT '',
              metodo_pago TEXT DEFAULT '',
              forma_pago TEXT DEFAULT '',
              tipo_cliente TEXT DEFAULT '',
              potencial TEXT DEFAULT '',
              etapa TEXT DEFAULT 'Prospecto nuevo',
              municipio TEXT DEFAULT '',
              localidad TEXT DEFAULT '',
              cabecera TEXT DEFAULT '',
              zona TEXT DEFAULT '',
              estado TEXT DEFAULT 'Nayarit',
              direccion TEXT DEFAULT '',
              contacto TEXT DEFAULT '',
              telefono TEXT DEFAULT '',
              whatsapp TEXT DEFAULT '',
              correo TEXT DEFAULT '',
              productos_interes TEXT DEFAULT '',
              consumo_mensual TEXT DEFAULT '',
              precio_actual_compra REAL,
              proveedor_actual TEXT DEFAULT '',
              precios_competencia TEXT DEFAULT '',
              observaciones TEXT DEFAULT '',
              info_enviada TEXT DEFAULT '',
              fecha_info_enviada TEXT DEFAULT '',
              created_at TEXT DEFAULT '',
              updated_at TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS interactions (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              client_id INTEGER NOT NULL,
              fecha TEXT NOT NULL,
              hora TEXT DEFAULT '',
              tipo_contacto TEXT DEFAULT '',
              comentarios TEXT DEFAULT '',
              productos_ofrecidos TEXT DEFAULT '',
              precio_ofrecido REAL,
              resultado TEXT DEFAULT '',
              proxima_accion TEXT DEFAULT '',
              fecha_seguimiento TEXT DEFAULT '',
              created_at TEXT DEFAULT '',
              FOREIGN KEY(client_id) REFERENCES clients(id)
            );
            CREATE TABLE IF NOT EXISTS imports (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              source TEXT,
              imported_at TEXT
            );
            CREATE TABLE IF NOT EXISTS quotes (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              client_id INTEGER NOT NULL,
              folio TEXT DEFAULT '',
              fecha TEXT DEFAULT '',
              cliente TEXT DEFAULT '',
              producto TEXT DEFAULT '',
              cantidad REAL DEFAULT 0,
              precio_unitario REAL DEFAULT 0,
              km_flete REAL DEFAULT 0,
              tarifa_km REAL DEFAULT 0,
              flete_base REAL DEFAULT 0,
              flete_total REAL DEFAULT 0,
              flete_unitario REAL DEFAULT 0,
              casetas REAL DEFAULT 0,
              total REAL DEFAULT 0,
              requiere_factura INTEGER DEFAULT 0,
              subtotal REAL DEFAULT 0,
              iva REAL DEFAULT 0,
              total_con_iva REAL DEFAULT 0,
              costo_individual REAL DEFAULT 0,
              status TEXT DEFAULT 'Elaborada',
              pdf TEXT DEFAULT '',
              created_at TEXT DEFAULT '',
              updated_at TEXT DEFAULT ''
            );
            """
        )
        ensure_business_tables(db)
        cols = {r["name"] for r in db.execute("PRAGMA table_info(clients)").fetchall()}
        if "localidad" not in cols:
            db.execute("ALTER TABLE clients ADD COLUMN localidad TEXT DEFAULT ''")
        if "info_enviada" not in cols:
            db.execute("ALTER TABLE clients ADD COLUMN info_enviada TEXT DEFAULT ''")
        if "fecha_info_enviada" not in cols:
            db.execute("ALTER TABLE clients ADD COLUMN fecha_info_enviada TEXT DEFAULT ''")
        if "created_by" not in cols:
            db.execute("ALTER TABLE clients ADD COLUMN created_by TEXT DEFAULT ''")
        if "created_by_name" not in cols:
            db.execute("ALTER TABLE clients ADD COLUMN created_by_name TEXT DEFAULT ''")
        if "assigned_to" not in cols:
            db.execute("ALTER TABLE clients ADD COLUMN assigned_to TEXT DEFAULT ''")
        fiscal_cols = {
            "tipo_persona_fiscal": "TEXT DEFAULT ''",
            "rfc": "TEXT DEFAULT ''",
            "regimen_fiscal": "TEXT DEFAULT ''",
            "uso_cfdi": "TEXT DEFAULT ''",
            "cp_fiscal": "TEXT DEFAULT ''",
            "domicilio_fiscal": "TEXT DEFAULT ''",
            "correo_facturacion": "TEXT DEFAULT ''",
            "metodo_pago": "TEXT DEFAULT ''",
            "forma_pago": "TEXT DEFAULT ''",
        }
        for col, definition in fiscal_cols.items():
            if col not in cols:
                db.execute(f"ALTER TABLE clients ADD COLUMN {col} {definition}")
        icols = {r["name"] for r in db.execute("PRAGMA table_info(interactions)").fetchall()}
        if "created_by" not in icols:
            db.execute("ALTER TABLE interactions ADD COLUMN created_by TEXT DEFAULT ''")
        if "created_by_name" not in icols:
            db.execute("ALTER TABLE interactions ADD COLUMN created_by_name TEXT DEFAULT ''")
        qcols = {r["name"] for r in db.execute("PRAGMA table_info(quotes)").fetchall()}
        quote_extra_cols = {
            "folio": "TEXT DEFAULT ''",
            "km_flete": "REAL DEFAULT 0",
            "tarifa_km": "REAL DEFAULT 0",
            "flete_base": "REAL DEFAULT 0",
            "flete_unitario": "REAL DEFAULT 0",
            "requiere_factura": "INTEGER DEFAULT 0",
            "subtotal": "REAL DEFAULT 0",
            "iva": "REAL DEFAULT 0",
            "total_con_iva": "REAL DEFAULT 0",
        }
        for col, definition in quote_extra_cols.items():
            if col not in qcols:
                db.execute(f"ALTER TABLE quotes ADD COLUMN {col} {definition}")
        if "created_by" not in qcols:
            db.execute("ALTER TABLE quotes ADD COLUMN created_by TEXT DEFAULT ''")
        if "created_by_name" not in qcols:
            db.execute("ALTER TABLE quotes ADD COLUMN created_by_name TEXT DEFAULT ''")
        db.execute("UPDATE clients SET created_by=?, created_by_name=? WHERE COALESCE(created_by,'')=''", (auth_user(), "Administrador BDP"))
        db.execute("UPDATE interactions SET created_by=?, created_by_name=? WHERE COALESCE(created_by,'')=''", (auth_user(), "Administrador BDP"))
        db.execute("UPDATE quotes SET created_by=?, created_by_name=? WHERE COALESCE(created_by,'')=''", (auth_user(), "Administrador BDP"))
        sync_regions(db)
        migrate_existing_quote_pdfs(db)
        db.commit()


def client_by_name_phone(db, name, phone):
    name = (name or "").strip()
    phone = (phone or "").strip()
    if phone:
        row = db.execute(
            "SELECT * FROM clients WHERE lower(nombre_comercial)=lower(?) OR telefono=? OR whatsapp=? LIMIT 1",
            (name, phone, phone),
        ).fetchone()
    else:
        row = db.execute("SELECT * FROM clients WHERE lower(nombre_comercial)=lower(?) LIMIT 1", (name,)).fetchone()
    return row


def upsert_client(db, data):
    municipio = data.get("municipio", "")
    catalog = NAYARIT_MUNICIPIOS.get(municipio, {})
    zona = data.get("zona") or catalog.get("zona", "")
    cabecera = data.get("cabecera") or catalog.get("cabecera", municipio)
    nombre = (data.get("nombre_comercial") or "Cliente sin nombre").strip()
    phone = (data.get("telefono") or data.get("whatsapp") or "").strip()
    existing = None
    if data.get("id"):
        existing = db.execute("SELECT * FROM clients WHERE id=?", (data.get("id"),)).fetchone()
    if not existing and not data.get("_force_new"):
        existing = client_by_name_phone(db, nombre, phone)
    payload = {
        "nombre_comercial": nombre,
        "razon_social": data.get("razon_social", ""),
        "tipo_persona_fiscal": data.get("tipo_persona_fiscal", ""),
        "rfc": data.get("rfc", ""),
        "regimen_fiscal": data.get("regimen_fiscal", ""),
        "uso_cfdi": data.get("uso_cfdi", ""),
        "cp_fiscal": data.get("cp_fiscal", ""),
        "domicilio_fiscal": data.get("domicilio_fiscal", ""),
        "correo_facturacion": data.get("correo_facturacion", ""),
        "metodo_pago": data.get("metodo_pago", ""),
        "forma_pago": data.get("forma_pago", ""),
        "tipo_cliente": data.get("tipo_cliente", ""),
        "potencial": data.get("potencial", ""),
        "etapa": data.get("etapa") or data.get("estatus") or "Prospecto nuevo",
        "municipio": municipio,
        "localidad": data.get("localidad", ""),
        "cabecera": cabecera,
        "zona": zona,
        "estado": data.get("estado") or "Nayarit",
        "direccion": data.get("direccion", ""),
        "contacto": data.get("contacto", ""),
        "telefono": data.get("telefono", ""),
        "whatsapp": data.get("whatsapp") or data.get("telefono", ""),
        "correo": data.get("correo", ""),
        "productos_interes": data.get("productos_interes", ""),
        "consumo_mensual": data.get("consumo_mensual", ""),
        "precio_actual_compra": data.get("precio_actual_compra"),
        "proveedor_actual": data.get("proveedor_actual", ""),
        "precios_competencia": data.get("precios_competencia", ""),
        "observaciones": data.get("observaciones", ""),
        "info_enviada": data.get("info_enviada", ""),
        "fecha_info_enviada": data.get("fecha_info_enviada", ""),
        "assigned_to": data.get("assigned_to", ""),
        "updated_at": now_iso(),
    }
    if existing and "assigned_to" not in data:
        payload["assigned_to"] = existing["assigned_to"] or ""
    if existing:
        # Actualización robusta: cuando el usuario edita una ficha, se deben guardar
        # TODOS los campos enviados por el formulario, incluyendo cambios en
        # tipo_cliente y campos que intencionalmente se dejan vacíos.
        # Antes se omitían valores vacíos, lo que podía dejar información vieja.
        fields = list(payload.keys())
        sets = ", ".join([f"{k}=?" for k in fields])
        db.execute(f"UPDATE clients SET {sets} WHERE id=?", [payload[k] for k in fields] + [existing["id"]])
        return existing["id"]
    payload["created_at"] = now_iso()
    payload["created_by"] = data.get("created_by", "")
    payload["created_by_name"] = data.get("created_by_name", "")
    if not payload["assigned_to"]:
        payload["assigned_to"] = payload["created_by"]
    keys = list(payload.keys())
    db.execute(
        f"INSERT INTO clients ({','.join(keys)}) VALUES ({','.join(['?'] * len(keys))})",
        [payload[k] for k in keys],
    )
    return db.execute("SELECT last_insert_rowid()").fetchone()[0]


def add_interaction(db, client_id, data):
    fecha = data.get("fecha") or date.today().isoformat()
    hora = data.get("hora") or datetime.now().strftime("%H:%M")
    db.execute(
        """
        INSERT INTO interactions
        (client_id, fecha, hora, tipo_contacto, comentarios, productos_ofrecidos, precio_ofrecido,
         resultado, proxima_accion, fecha_seguimiento, created_at, created_by, created_by_name)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            client_id,
            fecha,
            hora,
            data.get("tipo_contacto", ""),
            data.get("comentarios", ""),
            data.get("productos_ofrecidos", ""),
            data.get("precio_ofrecido") or None,
            data.get("resultado", ""),
            data.get("proxima_accion", ""),
            data.get("fecha_seguimiento", ""),
            now_iso(),
            data.get("created_by", ""),
            data.get("created_by_name", ""),
        ),
    )
    if data.get("resultado"):
        db.execute("UPDATE clients SET updated_at=? WHERE id=?", (now_iso(), client_id))


def update_interaction(db, interaction_id, data):
    row = db.execute("SELECT * FROM interactions WHERE id=?", (interaction_id,)).fetchone()
    if not row:
        return None
    fields = [
        "fecha",
        "hora",
        "tipo_contacto",
        "comentarios",
        "productos_ofrecidos",
        "precio_ofrecido",
        "resultado",
        "proxima_accion",
        "fecha_seguimiento",
    ]
    values = {}
    for field in fields:
        if field in data:
            values[field] = data.get(field) or (None if field == "precio_ofrecido" else "")
    if values:
        sets = ", ".join([f"{key}=?" for key in values])
        db.execute(f"UPDATE interactions SET {sets} WHERE id=?", list(values.values()) + [interaction_id])
        db.execute("UPDATE clients SET updated_at=? WHERE id=?", (now_iso(), row["client_id"]))
    if data.get("etapa"):
        db.execute("UPDATE clients SET etapa=?, updated_at=? WHERE id=?", (data["etapa"], now_iso(), row["client_id"]))
    return row["client_id"]


def import_excel(path_or_bytes, source="Excel"):
    if openpyxl is None:
        return {"imported": 0, "error": "openpyxl no está disponible"}
    wb = openpyxl.load_workbook(path_or_bytes, data_only=True)
    if "Captura_Visitas" not in wb.sheetnames:
        return {"imported": 0, "error": "No se encontró la hoja Captura_Visitas"}
    ws = wb["Captura_Visitas"]
    headers = [ws.cell(row=4, column=i).value for i in range(1, ws.max_column + 1)]
    count = 0
    with conn() as db:
        for row in ws.iter_rows(min_row=5, values_only=True):
            record = dict(zip(headers, row))
            name = record.get("Cliente / negocio")
            if not name:
                continue
            estatus = record.get("Estatus") or "Prospecto nuevo"
            etapa_map = {
                "Nuevo prospecto": "Prospecto nuevo",
                "Contactado": "Contactado",
                "Cotización enviada": "Cotización enviada",
                "Seguimiento": "Negociación",
                "Distribuidor captado": "Cliente activo",
                "Cliente activo": "Cliente activo",
                "No interesado": "Cliente perdido",
            }
            potencial_map = {"Alto": "A", "Medio": "B", "Bajo": "C", "A": "A", "B": "B", "C": "C"}
            client_id = upsert_client(
                db,
                {
                    "nombre_comercial": name,
                    "contacto": record.get("Contacto") or "",
                    "telefono": str(record.get("Teléfono") or ""),
                    "whatsapp": str(record.get("Teléfono") or ""),
                    "tipo_cliente": record.get("Tipo cliente") or "",
                    "municipio": record.get("Municipio") or "",
                    "cabecera": record.get("Cabecera municipal") or "",
                    "zona": record.get("Zona") or "",
                    "productos_interes": record.get("Producto") or "",
                    "precio_actual_compra": record.get("Precio unitario planta"),
                    "observaciones": record.get("Observaciones") or "",
                    "potencial": potencial_map.get(str(record.get("Potencial") or ""), str(record.get("Potencial") or "")),
                    "etapa": etapa_map.get(str(estatus), str(estatus)),
                },
            )
            add_interaction(
                db,
                client_id,
                {
                    "fecha": normalize_date(record.get("Fecha visita")),
                    "tipo_contacto": "Visita",
                    "comentarios": record.get("Observaciones") or record.get("Próxima acción") or "",
                    "productos_ofrecidos": record.get("Producto") or "",
                    "precio_ofrecido": record.get("Costo individual ofrecido") or record.get("Precio ofrecido unitario"),
                    "resultado": str(estatus),
                    "proxima_accion": record.get("Próxima acción") or "",
                    "fecha_seguimiento": normalize_date(record.get("Fecha seguimiento")),
                },
            )
            count += 1
        db.execute("INSERT INTO imports (source, imported_at) VALUES (?, ?)", (source, now_iso()))
        db.commit()
    return {"imported": count}


def seed_initial_data():
    with conn() as db:
        existing = db.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
    if existing == 0 and SOURCE_XLSX.exists():
        import_excel(SOURCE_XLSX, "Base inicial")


def auto_backup():
    if not DB_PATH.exists():
        return
    stamp = date.today().isoformat()
    target = BACKUP_DIR / f"bdp_crm_backup_{stamp}.sqlite"
    if not target.exists():
        shutil.copy2(DB_PATH, target)


def rows_to_dicts(rows):
    return [dict(r) for r in rows]


def export_xlsx():
    if openpyxl is None:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    headers = [
        "ID", "Nombre comercial", "Razón social", "Tipo cliente", "Potencial", "Etapa", "Zona", "Municipio",
        "Localidad", "Estado", "Dirección", "Contacto", "Teléfono", "WhatsApp", "Correo", "Productos interés",
        "Consumo mensual", "Precio actual compra", "Proveedor actual", "Precios competencia", "Observaciones",
        "Info enviada", "Fecha info enviada",
    ]
    ws.append(headers)
    with conn() as db:
        for r in db.execute("SELECT * FROM clients ORDER BY nombre_comercial"):
            ws.append([
                r["id"], r["nombre_comercial"], r["razon_social"], r["tipo_cliente"], r["potencial"], r["etapa"],
                r["zona"], r["municipio"], r["localidad"], r["estado"], r["direccion"], r["contacto"], r["telefono"], r["whatsapp"],
                r["correo"], r["productos_interes"], r["consumo_mensual"], r["precio_actual_compra"],
                r["proveedor_actual"], r["precios_competencia"], r["observaciones"], r["info_enviada"], r["fecha_info_enviada"],
            ])
    ws2 = wb.create_sheet("Historial")
    ws2.append(["Cliente ID", "Cliente", "Fecha", "Hora", "Tipo contacto", "Comentarios", "Productos ofrecidos", "Precio ofrecido", "Resultado", "Próxima acción", "Fecha seguimiento"])
    with conn() as db:
        q = """
        SELECT i.*, c.nombre_comercial
        FROM interactions i JOIN clients c ON c.id=i.client_id
        ORDER BY i.fecha DESC, i.hora DESC
        """
        for r in db.execute(q):
            ws2.append([r["client_id"], r["nombre_comercial"], r["fecha"], r["hora"], r["tipo_contacto"], r["comentarios"], r["productos_ofrecidos"], r["precio_ofrecido"], r["resultado"], r["proxima_accion"], r["fecha_seguimiento"]])
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col[:80])
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 38)
        sheet.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def money(value):
    try:
        return f"${float(value):,.2f}"
    except Exception:
        return "$0.00"


def clean_code_text(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return "".join(ch for ch in text.upper() if ch.isalnum() or ch.isspace())


def product_code(product):
    text = clean_code_text(product)
    if "10" in text and "14" in text and "28" in text:
        return "B1028"
    if "12" in text and "20" in text and "40" in text:
        return "B1240"
    if "14" in text and "20" in text and "40" in text:
        return "B1440"
    if "BOVEDILLA" in text:
        return "BOV"
    if "ADOQUIN" in text or "ADOQUIN" in text:
        return "ADO"
    initials = "".join(word[0] for word in text.split() if word)
    return (initials[:4] or "MAT")


def build_quote_folio(db, fecha, producto):
    try:
        quote_date = datetime.strptime(fecha or "", "%Y-%m-%d").date()
    except Exception:
        quote_date = date.today()
    prefix = f"COT-{quote_date.strftime('%y%m%d')}-{product_code(producto)}"
    count = db.execute("SELECT COUNT(*) FROM quotes WHERE folio LIKE ?", (prefix + "-%",)).fetchone()[0] + 1
    return f"{prefix}-{count:03d}"


def _words_0_999(number):
    units = ["", "uno", "dos", "tres", "cuatro", "cinco", "seis", "siete", "ocho", "nueve"]
    teens = {
        10: "diez", 11: "once", 12: "doce", 13: "trece", 14: "catorce", 15: "quince",
        16: "dieciseis", 17: "diecisiete", 18: "dieciocho", 19: "diecinueve",
    }
    tens = ["", "", "veinte", "treinta", "cuarenta", "cincuenta", "sesenta", "setenta", "ochenta", "noventa"]
    hundreds = ["", "ciento", "doscientos", "trescientos", "cuatrocientos", "quinientos", "seiscientos", "setecientos", "ochocientos", "novecientos"]
    number = int(number)
    if number == 0:
        return ""
    if number == 100:
        return "cien"
    if number < 10:
        return units[number]
    if number < 20:
        return teens[number]
    if number < 30:
        special = {20: "veinte", 21: "veintiuno", 22: "veintidos", 23: "veintitres", 24: "veinticuatro", 25: "veinticinco", 26: "veintiseis", 27: "veintisiete", 28: "veintiocho", 29: "veintinueve"}
        return special[number]
    if number < 100:
        ten, unit = divmod(number, 10)
        return tens[ten] if unit == 0 else f"{tens[ten]} y {units[unit]}"
    hundred, rest = divmod(number, 100)
    return hundreds[hundred] if rest == 0 else f"{hundreds[hundred]} {_words_0_999(rest)}"


def number_to_words_es(number):
    number = int(number)
    if number == 0:
        return "cero"
    if number < 1000:
        return _words_0_999(number)
    if number < 1_000_000:
        thousands, rest = divmod(number, 1000)
        prefix = "mil" if thousands == 1 else f"{_words_0_999(thousands)} mil"
        return prefix if rest == 0 else f"{prefix} {_words_0_999(rest)}"
    millions, rest = divmod(number, 1_000_000)
    prefix = "un millon" if millions == 1 else f"{number_to_words_es(millions)} millones"
    return prefix if rest == 0 else f"{prefix} {number_to_words_es(rest)}"


def amount_words_for_pesos(pesos):
    words = number_to_words_es(pesos)
    replacements = (
        ("veintiuno", "veintiun"),
        (" y uno", " y un"),
        (" uno", " un"),
    )
    if words == "uno":
        return "un"
    for old, new in replacements:
        if words.endswith(old):
            return words[: -len(old)] + new
    return words


def amount_to_words_mxn(value):
    amount = round(float(value or 0), 2)
    pesos = int(amount)
    cents = int(round((amount - pesos) * 100))
    if cents == 100:
        pesos += 1
        cents = 0
    currency = "PESO" if pesos == 1 else "PESOS"
    return f"SON: {amount_words_for_pesos(pesos).upper()} {currency} {cents:02d}/100 M.N."


def generate_quote_pdf(client, data):
    if SimpleDocTemplate is None:
        return None
    safe_name = "".join(ch for ch in client["nombre_comercial"] if ch.isalnum() or ch in (" ", "_", "-")).strip().replace(" ", "_")
    filename = f"{data.get('folio') or 'cotizacion'}_{safe_name or 'cliente'}.pdf"
    path = QUOTE_DIR / filename
    logo = APP_DIR / "assets" / "logo.png"
    doc = SimpleDocTemplate(str(path), pagesize=letter, rightMargin=28, leftMargin=28, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    styles["Title"].fontSize = 15
    styles["Title"].leading = 17
    styles["Heading2"].fontSize = 12
    styles["Heading2"].leading = 14
    styles["Heading3"].fontSize = 9
    styles["Heading3"].leading = 11
    styles["Normal"].fontSize = 8
    styles["Normal"].leading = 10
    story = []
    if logo.exists():
        story.append(Image(str(logo), width=105, height=52, kind="proportional"))
    story.append(Paragraph("<b>Bazar de Prefabricados</b>", styles["Title"]))
    story.append(Paragraph("Cotizacion comercial - Soluciones en concreto", styles["Normal"]))
    story.append(Paragraph(f"<b>Folio:</b> {data.get('folio') or '-'} | <b>Fecha:</b> {data.get('fecha') or date.today().isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"<b>Cliente:</b> {client['nombre_comercial']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Contacto:</b> {client['contacto'] or '-'} | <b>WhatsApp:</b> {client['whatsapp'] or client['telefono'] or '-'}", styles["Normal"]))
    story.append(Paragraph(f"<b>Municipio:</b> {client['municipio'] or '-'} | <b>Region:</b> {client['zona'] or '-'}", styles["Normal"]))
    story.append(Paragraph(f"<b>Elaboro:</b> {data.get('created_by_name') or data.get('created_by') or '-'}", styles["Normal"]))
    story.append(Spacer(1, 8))
    rows = [
        ["Concepto", "Importe / dato"],
        ["Folio", data.get("folio") or "-"],
        ["Producto", data.get("producto") or data.get("productos_ofrecidos") or "-"],
        ["Cantidad de piezas", data.get("cantidad") or "-"],
        ["Precio unitario planta", money(data.get("precio_unitario"))],
        ["Kilometros flete", data.get("km_flete") or "0"],
        ["Tarifa flete por km", money(data.get("tarifa_km"))],
        ["Flete base", money(data.get("flete_base"))],
        ["Casetas de cobro", money(data.get("casetas"))],
        ["Flete total", money(data.get("flete_total"))],
        ["Flete por pieza", money(data.get("flete_unitario"))],
        ["Costo individual con flete", money(data.get("precio_ofrecido"))],
        ["Total cotizado", money(data.get("total"))],
        ["Requiere factura", "Si" if data.get("requiere_factura") else "No"],
    ]
    if data.get("requiere_factura"):
        rows.extend([
            ["Subtotal", money(data.get("subtotal"))],
            ["IVA incluido 16%", money(data.get("iva"))],
            ["Total final con IVA incluido", money(data.get("total_con_iva"))],
        ])
    table = Table(rows, colWidths=[190, 330])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#c99740")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f7f3e8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("LEADING", (0, 0), (-1, -1), 8),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    final_total = float(data.get("total_con_iva") or data.get("total") or 0)
    story.append(Spacer(1, 5))
    story.append(Paragraph(f"<b>COSTO FINAL: {money(final_total)}</b>", styles["Heading2"]))
    story.append(Paragraph(f"<b>{amount_to_words_mxn(final_total)}</b>", styles["Heading3"]))
    story.append(Spacer(1, 8))
    fiscal_rows = [
        ["Dato fiscal", "Informacion capturada"],
        ["Razon social", client["razon_social"] or "-"],
        ["Tipo persona", client["tipo_persona_fiscal"] or "-"],
        ["RFC", client["rfc"] or "-"],
        ["Regimen fiscal", client["regimen_fiscal"] or "-"],
        ["Uso CFDI", client["uso_cfdi"] or "-"],
        ["Codigo postal fiscal", client["cp_fiscal"] or "-"],
        ["Domicilio fiscal", client["domicilio_fiscal"] or "-"],
        ["Correo facturacion", client["correo_facturacion"] or client["correo"] or "-"],
        ["Metodo de pago", client["metodo_pago"] or "-"],
        ["Forma de pago", client["forma_pago"] or "-"],
    ]
    fiscal_table = Table(fiscal_rows, colWidths=[175, 345])
    fiscal_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f1b14")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#aaaaaa")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#fbf8ef")),
        ("FONTSIZE", (0, 0), (-1, -1), 6.5),
        ("LEADING", (0, 0), (-1, -1), 7.5),
        ("PADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(Paragraph("<b>Datos fiscales para facturacion</b>", styles["Heading3"]))
    story.append(fiscal_table)
    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>En caso de requerir factura favor de confirmar sus datos fiscales.</b>", styles["Normal"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph("<b>Observaciones:</b>", styles["Heading3"]))
    story.append(Paragraph(data.get("comentarios", "Precio sujeto a confirmacion de volumen, ruta y disponibilidad."), styles["Normal"]))
    story.append(Spacer(1, 5))
    story.append(Paragraph("Contacto: 311 105 1368 | bdprefabricados@hotmail.com | Tepic, Nayarit", styles["Normal"]))
    doc.build(story)
    return filename


def quote_report_rows(status="", month="", user=None):
    scope, scope_params = quote_owner_sql(user, "q")
    where = [scope]
    params = list(scope_params)
    if status and status != "Todas":
        where.append("q.status=?")
        params.append(status)
    if month:
        where.append("q.fecha LIKE ?")
        params.append(f"{month}%")
    where_sql = "WHERE " + " AND ".join(where)
    with conn() as db:
        rows = db.execute(
            f"""
            SELECT q.*, c.municipio, c.zona, c.contacto, c.whatsapp, c.telefono, c.correo, c.correo_facturacion
            FROM quotes q
            LEFT JOIN clients c ON c.id=q.client_id
            {where_sql}
            ORDER BY q.fecha DESC, q.id DESC
            """,
            params,
        ).fetchall()
    return rows_to_dicts(rows)


def export_quotes_report_xlsx(rows, status="", month=""):
    if openpyxl is None:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizaciones"
    ws.append(["Reporte de cotizaciones", f"Estatus: {status or 'Todas'}", f"Periodo: {month or 'Todos'}"])
    ws.append([])
    headers = ["Folio", "Fecha", "Cliente", "Municipio", "Zona", "Producto", "Cantidad", "Precio unitario", "Flete", "Casetas", "Total IVA incluido", "Estatus", "Elaboro", "PDF"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row.get("folio") or "",
            row.get("fecha") or "",
            row.get("cliente") or "",
            row.get("municipio") or "",
            row.get("zona") or "",
            row.get("producto") or "",
            row.get("cantidad") or 0,
            row.get("precio_unitario") or 0,
            row.get("flete_total") or 0,
            row.get("casetas") or 0,
            row.get("total_con_iva") or row.get("total") or 0,
            row.get("status") or "",
            row.get("created_by_name") or row.get("created_by") or "",
            row.get("pdf") or "",
        ])
    for col in range(1, len(headers) + 1):
        ws.cell(row=3, column=col).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_quotes_report_pdf(rows, status="", month=""):
    if SimpleDocTemplate is None:
        return None
    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=letter, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    styles["Title"].fontSize = 14
    styles["Normal"].fontSize = 7
    styles["Normal"].leading = 8
    story = [
        Paragraph("<b>Bazar de Prefabricados</b>", styles["Title"]),
        Paragraph(f"Reporte de cotizaciones | Estatus: {status or 'Todas'} | Periodo: {month or 'Todos'}", styles["Normal"]),
        Spacer(1, 8),
    ]
    table_rows = [["Folio", "Fecha", "Cliente", "Municipio", "Producto", "Total", "Estatus"]]
    for row in rows:
        table_rows.append([
            row.get("folio") or "",
            row.get("fecha") or "",
            row.get("cliente") or "",
            row.get("municipio") or "",
            row.get("producto") or "",
            money(row.get("total_con_iva") or row.get("total") or 0),
            row.get("status") or "",
        ])
    table = Table(table_rows, colWidths=[82, 58, 105, 80, 105, 62, 72])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#c99740")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#999999")),
        ("FONTSIZE", (0, 0), (-1, -1), 6.3),
        ("LEADING", (0, 0), (-1, -1), 7),
        ("PADDING", (0, 0), (-1, -1), 2.5),
    ]))
    story.append(table)
    doc.build(story)
    return bio.getvalue()


def send_quote_email(client, data, pdf_path):
    with conn() as db:
        settings = {r["key"]: r["value"] for r in db.execute("SELECT key,value FROM app_settings").fetchall()}
    host = os.environ.get("SMTP_HOST") or settings.get("smtp_host") or ""
    port = int(os.environ.get("SMTP_PORT") or settings.get("smtp_port") or "587")
    user = os.environ.get("SMTP_USER") or settings.get("smtp_user") or ""
    password = os.environ.get("SMTP_PASSWORD") or settings.get("smtp_password") or ""
    from_email = os.environ.get("SMTP_FROM_EMAIL") or settings.get("smtp_from_email") or user
    from_name = settings.get("smtp_from_name") or "BAZAR DE PREFABRICADOS"
    admin_email = settings.get("quote_admin_email") or "bdprefabricados@hotmail.com"
    client_email = client["correo_facturacion"] or client["correo"] or ""
    recipients = [email for email in [client_email, admin_email] if email]
    if not recipients:
        return {"sent": False, "error": "El cliente no tiene correo y no hay correo administrador configurado."}
    if not host or not from_email:
        return {"sent": False, "error": "Falta configurar correo SMTP en Configuracion."}
    msg = EmailMessage()
    msg["Subject"] = f"Anexo cotizacion solicitada {data.get('folio') or ''}".strip()
    msg["From"] = f"{from_name} <{from_email}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content("Anexo cotizacion solicitada.\n\nBAZAR DE PREFABRICADOS\nSoluciones en concreto")
    with open(pdf_path, "rb") as fh:
        msg.add_attachment(fh.read(), maintype="application", subtype="pdf", filename=Path(pdf_path).name)
    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port, timeout=20) as smtp:
                if user:
                    smtp.login(user, password)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=20) as smtp:
                smtp.starttls()
                if user:
                    smtp.login(user, password)
                smtp.send_message(msg)
        return {"sent": True, "recipients": recipients}
    except Exception as exc:
        return {"sent": False, "error": str(exc)}


def auth_user():
    return os.environ.get("CRM_USER", "admin")


def auth_password():
    return os.environ.get("CRM_PASSWORD", "1234")


def auth_secret():
    return os.environ.get("SESSION_SECRET", auth_password() + "_crm_bdp_2026")


def hash_password(password, salt=None):
    salt = salt or hashlib.sha256(os.urandom(16)).hexdigest()[:16]
    digest = hashlib.pbkdf2_hmac("sha256", str(password).encode("utf-8"), salt.encode("utf-8"), 120000).hex()
    return f"{salt}${digest}"


def verify_password(password, stored):
    if not stored or "$" not in stored:
        return False
    salt, expected = stored.split("$", 1)
    actual = hash_password(password, salt).split("$", 1)[1]
    return hmac.compare_digest(actual, expected)


def session_token(username, password_hash):
    raw = f"{username}:{password_hash}".encode("utf-8")
    return hmac.new(auth_secret().encode("utf-8"), raw, hashlib.sha256).hexdigest()


def public_user(row):
    if not row:
        return None
    return {
        "id": row["id"],
        "username": row["username"],
        "full_name": row["full_name"],
        "role": row["role"],
        "active": row["active"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def is_admin(user):
    return bool(user and user["role"] == "Administrador")


def owner_sql(user, alias=""):
    """Return SQL and parameters limiting commercial data to the signed-in seller."""
    if is_admin(user):
        return "1=1", []
    prefix = f"{alias}." if alias else ""
    username = user["username"] if user else ""
    return f"({prefix}created_by=? OR {prefix}assigned_to=?)", [username, username]


def quote_owner_sql(user, alias=""):
    if is_admin(user):
        return "1=1", []
    prefix = f"{alias}." if alias else ""
    username = user["username"] if user else ""
    return f"{prefix}created_by=?", [username]


def can_access_client(db, user, client_id):
    clause, params = owner_sql(user)
    return db.execute(f"SELECT id FROM clients WHERE id=? AND {clause}", [client_id] + params).fetchone() is not None


def can_access_quote(db, user, quote_id):
    clause, params = quote_owner_sql(user)
    return db.execute(f"SELECT id FROM quotes WHERE id=? AND {clause}", [quote_id] + params).fetchone() is not None


def login_page(error=""):
    message = f"<div class='error'>{error}</div>" if error else ""
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Acceso CRM BDP</title>
  <style>
    body{{margin:0;min-height:100vh;display:grid;place-items:center;background:#111;color:#f6f0df;font-family:Arial,sans-serif}}
    .box{{width:min(420px,92vw);background:#1b1b1b;border:1px solid #8f7b52;border-radius:12px;padding:28px;box-shadow:0 22px 60px rgba(0,0,0,.45)}}
    h1{{margin:0 0 8px;color:#c9a15a}} p{{color:#d5d0c4}}
    label{{display:block;margin:14px 0 6px;font-size:13px;color:#c9a15a;font-weight:bold}}
    input{{width:100%;box-sizing:border-box;padding:13px;border-radius:8px;border:1px solid #444;background:#0d0d0d;color:white;font-size:16px}}
    button{{width:100%;margin-top:18px;padding:13px;border:0;border-radius:8px;background:#c9a15a;color:#111;font-size:16px;font-weight:800;cursor:pointer}}
    .error{{background:#3a1717;color:#ffdede;border:1px solid #9a4444;padding:10px;border-radius:8px;margin:12px 0}}
  </style>
</head>
<body>
  <form class="box" method="post" action="/api/login">
    <h1>CRM BDP 2026</h1>
    <p>Ingresa usuario y contraseña para acceder.</p>
    {message}
    <label>Usuario</label>
    <input name="user" autocomplete="username" required>
    <label>Contraseña</label>
    <input name="password" type="password" autocomplete="current-password" required>
    <button>Entrar</button>
  </form>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    def _send(self, status=200, content_type="application/json", body=b"", headers=None):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store")
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        if isinstance(body, str):
            body = body.encode("utf-8")
        self.wfile.write(body)

    def json(self, obj, status=200):
        self._send(status, "application/json; charset=utf-8", json.dumps(obj, ensure_ascii=False).encode("utf-8"))

    def current_user(self):
        cookie = self.headers.get("Cookie", "")
        token = ""
        for part in cookie.split(";"):
            part = part.strip()
            if part.startswith("crm_session="):
                token = part.split("=", 1)[1]
                break
        if "|" not in token:
            return None
        username, signature = token.split("|", 1)
        with conn() as db:
            row = db.execute("SELECT * FROM users WHERE username=? AND active=1", (username,)).fetchone()
        if not row:
            return None
        expected = session_token(row["username"], row["password_hash"])
        if not hmac.compare_digest(signature, expected):
            return None
        return row

    def is_authenticated(self):
        return self.current_user() is not None

    def require_auth(self, path):
        if self.current_user():
            return True
        if path.startswith("/api/"):
            self.json({"error": "No autorizado"}, 401)
        else:
            self._send(200, "text/html; charset=utf-8", login_page())
        return False

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        qs = parse_qs(parsed.query)
        if path == "/login":
            return self._send(200, "text/html; charset=utf-8", login_page())
        if path == "/api/logout":
            return self._send(
                303,
                "text/plain; charset=utf-8",
                "OK",
                {
                    "Set-Cookie": "crm_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax",
                    "Location": "/login",
                },
            )
        if path.startswith("/assets/"):
            filename = Path(path).name
            target = APP_DIR / "assets" / filename
            if target.exists() and target.is_file():
                ext = target.suffix.lower()
                content_type = "application/octet-stream"
                if ext == ".png":
                    content_type = "image/png"
                elif ext in (".jpg", ".jpeg"):
                    content_type = "image/jpeg"
                elif ext == ".gif":
                    content_type = "image/gif"
                elif ext == ".svg":
                    content_type = "image/svg+xml"
                elif ext == ".css":
                    content_type = "text/css; charset=utf-8"
                elif ext == ".js":
                    content_type = "application/javascript; charset=utf-8"
                return self._send(200, content_type, target.read_bytes())
            return self.json({"error": "Asset no encontrado"}, 404)
        if not self.require_auth(path):
            return
        if path in ("/", "/index.html"):
            return self._send(200, "text/html; charset=utf-8", (APP_DIR / "index.html").read_bytes())
        if path.startswith("/assets/"):
            filename = Path(path).name
            target = APP_DIR / "assets" / filename
            if target.exists() and target.is_file():
                ext = target.suffix.lower()
                content_type = "application/octet-stream"
                if ext == ".png":
                    content_type = "image/png"
                elif ext in (".jpg", ".jpeg"):
                    content_type = "image/jpeg"
                elif ext == ".gif":
                    content_type = "image/gif"
                elif ext == ".svg":
                    content_type = "image/svg+xml"
                elif ext == ".css":
                    content_type = "text/css; charset=utf-8"
                elif ext == ".js":
                    content_type = "application/javascript; charset=utf-8"
                return self._send(200, content_type, target.read_bytes())
            return self.json({"error": "Asset no encontrado"}, 404)
        if path.startswith("/cotizaciones_pdf/"):
            filename = Path(path).name
            target = QUOTE_DIR / filename
            with conn() as db:
                quote = db.execute("SELECT id FROM quotes WHERE pdf LIKE ?", (f"%/{filename}",)).fetchone()
                allowed = quote and can_access_quote(db, self.current_user(), quote["id"])
            if allowed and target.exists() and target.is_file():
                return self._send(200, "application/pdf", target.read_bytes())
            return self.json({"error": "PDF no encontrado o sin permiso"}, 404)
        if path.startswith("/cotizaciones_logradas/"):
            filename = Path(path).name
            target = QUOTE_WON_DIR / filename
            with conn() as db:
                quote = db.execute("SELECT id FROM quotes WHERE pdf LIKE ?", (f"%/{filename}",)).fetchone()
                allowed = quote and can_access_quote(db, self.current_user(), quote["id"])
            if allowed and target.exists() and target.is_file():
                return self._send(200, "application/pdf", target.read_bytes())
            return self.json({"error": "PDF no encontrado"}, 404)
        if path == "/api/options":
            
            with conn() as db:
                cfg = get_settings_payload(db)
            return self.json({"municipios": NAYARIT_MUNICIPIOS, "productos": cfg["productos"], "tarifas": cfg["tarifas"], "regiones": REGIONES, "sat": SAT_CATALOGS})
        if path == "/api/me":
            return self.json({"user": public_user(self.current_user())})
        if path == "/api/users":
            user = self.current_user()
            if not user or user["role"] != "Administrador":
                return self.json({"error": "Solo administrador"}, 403)
            with conn() as db:
                rows = db.execute("SELECT id, username, full_name, role, active, created_at, updated_at FROM users ORDER BY active DESC, full_name, username").fetchall()
            return self.json(rows_to_dicts(rows))
        if path == "/api/clients":
            q = (qs.get("q", [""])[0] or "").lower()
            scope, scope_params = owner_sql(self.current_user())
            with conn() as db:
                if q:
                    like = f"%{q}%"
                    rows = db.execute(
                        f"""
                        SELECT * FROM clients
                        WHERE ({scope}) AND (lower(nombre_comercial) LIKE ? OR lower(contacto) LIKE ? OR lower(municipio) LIKE ?
                           OR lower(tipo_cliente) LIKE ? OR lower(etapa) LIKE ? OR lower(potencial) LIKE ?
                           OR telefono LIKE ? OR whatsapp LIKE ?)
                        ORDER BY updated_at DESC, nombre_comercial
                        """,
                        scope_params + [like, like, like, like, like, like, like, like],
                    ).fetchall()
                else:
                    rows = db.execute(
                        f"""
                        SELECT * FROM clients WHERE {scope}
                        ORDER BY
                          CASE zona
                            WHEN 'Costa Sur' THEN 1
                            WHEN 'Costa Centro y Valles del Norte' THEN 2
                            WHEN 'Costa Norte y Frontera' THEN 3
                            WHEN 'Centro y Sierra Centro' THEN 4
                            WHEN 'Sierra y Sur' THEN 5
                            ELSE 6
                          END,
                          nombre_comercial COLLATE NOCASE
                        """, scope_params
                    ).fetchall()
            return self.json(rows_to_dicts(rows))
        if path == "/api/client":
            cid = qs.get("id", [""])[0]
            with conn() as db:
                if not can_access_client(db, self.current_user(), cid):
                    return self.json({"error": "Cliente no encontrado o sin permiso"}, 404)
                client = db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
                if not client:
                    return self.json({"error": "Cliente no encontrado"}, 404)
                hist = db.execute("SELECT * FROM interactions WHERE client_id=? ORDER BY fecha DESC, hora DESC, id DESC", (cid,)).fetchall()
            return self.json({"client": dict(client), "history": rows_to_dicts(hist)})
        if path == "/api/dashboard":
            today = date.today()
            overdue = today.isoformat()
            thirty = (today - timedelta(days=30)).isoformat()
            month_prefix = today.strftime("%Y-%m")
            client_scope, client_params = owner_sql(self.current_user())
            c_scope, c_params = owner_sql(self.current_user(), "c")
            quote_scope, quote_params = quote_owner_sql(self.current_user())
            with conn() as db:
                total = db.execute(f"SELECT COUNT(*) FROM clients WHERE {client_scope}", client_params).fetchone()[0]
                active = db.execute(f"SELECT COUNT(*) FROM clients WHERE ({client_scope}) AND etapa='Cliente activo'", client_params).fetchone()[0]
                pending_quotes = db.execute(f"SELECT COUNT(*) FROM clients WHERE ({client_scope}) AND etapa='Cotización enviada'", client_params).fetchone()[0]
                quotes_month = db.execute(f"SELECT COUNT(*) FROM quotes WHERE ({quote_scope}) AND fecha LIKE ?", quote_params + [month_prefix + "%"]).fetchone()[0]
                quotes_won = db.execute(f"SELECT COUNT(*) FROM quotes WHERE ({quote_scope}) AND status='Lograda'", quote_params).fetchone()[0]
                quotes_total_amount = db.execute(f"SELECT COALESCE(SUM(total),0) FROM quotes WHERE {quote_scope}", quote_params).fetchone()[0]
                quotes_won_amount = db.execute(f"SELECT COALESCE(SUM(total),0) FROM quotes WHERE ({quote_scope}) AND status='Lograda'", quote_params).fetchone()[0]
                follows = db.execute(f"SELECT COUNT(*) FROM interactions i JOIN clients c ON c.id=i.client_id WHERE ({c_scope}) AND i.fecha_seguimiento<>'' AND i.fecha_seguimiento<=?", c_params + [overdue]).fetchone()[0]
                by_stage = rows_to_dicts(db.execute(f"SELECT etapa, COUNT(*) total FROM clients WHERE {client_scope} GROUP BY etapa ORDER BY total DESC", client_params).fetchall())
                by_mun = rows_to_dicts(db.execute(f"SELECT municipio, zona, COUNT(*) total FROM clients WHERE {client_scope} GROUP BY municipio, zona ORDER BY total DESC", client_params).fetchall())
                by_zone_clients = rows_to_dicts(db.execute(
                    f"""
                    SELECT id, nombre_comercial, contacto, municipio, localidad, zona, etapa, potencial, tipo_cliente, whatsapp, telefono, created_at,
                           created_by, created_by_name, assigned_to
                    FROM clients WHERE {client_scope}
                    ORDER BY
                      CASE zona
                        WHEN 'Costa Sur' THEN 1
                        WHEN 'Costa Centro y Valles del Norte' THEN 2
                        WHEN 'Costa Norte y Frontera' THEN 3
                        WHEN 'Centro y Sierra Centro' THEN 4
                        WHEN 'Sierra y Sur' THEN 5
                        ELSE 6
                      END,
                      municipio,
                      nombre_comercial
                    """, client_params
                ).fetchall())
                alerts = rows_to_dicts(db.execute(
                    f"""
                    SELECT c.id, c.nombre_comercial, c.municipio, c.localidad, c.zona, c.etapa, c.info_enviada,
                           MAX(i.fecha) ultima_fecha,
                           MIN(CASE WHEN i.fecha_seguimiento<>'' THEN i.fecha_seguimiento END) seguimiento
                    FROM clients c LEFT JOIN interactions i ON i.client_id=c.id
                    WHERE {c_scope}
                    GROUP BY c.id
                    HAVING (seguimiento IS NOT NULL AND seguimiento<=?)
                       OR (ultima_fecha IS NOT NULL AND ultima_fecha<=?)
                       OR c.etapa='Cotización enviada'
                       OR c.info_enviada=''
                    ORDER BY seguimiento ASC
                    LIMIT 60
                    """,
                    c_params + [overdue, thirty],
                ).fetchall())
                volume = db.execute(f"SELECT COUNT(*) FROM clients WHERE ({client_scope}) AND potencial='A'", client_params).fetchone()[0]
                cfg = get_settings_payload(db)
                month_now = today.month
                sales_rows = cfg["monthlySales"]
                current_sales = next((r for r in sales_rows if int(r["month"]) == month_now), {"amount": 0, "goal": 0})
                annual_goal = float(cfg["settings"].get("annual_goal") or 0)
                annual_sales = sum(float(r.get("amount") or 0) for r in sales_rows)
            return self.json({"total": total, "active": active, "pendingQuotes": pending_quotes, "quotesMonth": quotes_month, "quotesWon": quotes_won, "quotesTotalAmount": quotes_total_amount, "quotesWonAmount": quotes_won_amount, "followUps": follows, "byStage": by_stage, "byMunicipio": by_mun, "byZoneClients": by_zone_clients, "alerts": alerts, "altoPotencial": volume, "settings": cfg["settings"], "monthlySales": sales_rows, "currentSales": current_sales, "annualSales": annual_sales, "annualGoal": annual_goal, "months": MESES_ES})
        if path == "/api/quotes":
            scope, scope_params = quote_owner_sql(self.current_user(), "q")
            with conn() as db:
                rows = db.execute(f"""
                    SELECT q.*, c.municipio, c.zona, c.contacto, c.whatsapp, c.telefono
                    FROM quotes q LEFT JOIN clients c ON c.id=q.client_id
                    WHERE {scope}
                    ORDER BY q.created_at DESC, q.id DESC
                """, scope_params).fetchall()
            return self.json(rows_to_dicts(rows))
        if path == "/api/settings":
            year = qs.get("year", [str(date.today().year)])[0]
            with conn() as db:
                return self.json(get_settings_payload(db, year))
        if path in ("/api/quotes_report.xlsx", "/api/quotes_report.pdf"):
            status = qs.get("status", [""])[0]
            month = qs.get("month", [""])[0]
            rows = quote_report_rows(status, month, self.current_user())
            if path.endswith(".xlsx"):
                data = export_quotes_report_xlsx(rows, status, month)
                if data is None:
                    return self.json({"error": "No se pudo generar Excel"}, 500)
                return self._send(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", data, {
                    "Content-Disposition": f"attachment; filename=reporte_cotizaciones_{month or 'todo'}_{status or 'todas'}.xlsx"
                })
            data = export_quotes_report_pdf(rows, status, month)
            if data is None:
                return self.json({"error": "No se pudo generar PDF"}, 500)
            return self._send(200, "application/pdf", data, {
                "Content-Disposition": f"attachment; filename=reporte_cotizaciones_{month or 'todo'}_{status or 'todas'}.pdf"
            })
        if path == "/api/export.xlsx":
            if not is_admin(self.current_user()):
                return self.json({"error": "Solo el administrador puede exportar toda la base"}, 403)
            data = export_xlsx()
            if data is None:
                return self.json({"error": "No se pudo exportar a Excel"}, 500)
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename=bdp_crm_export.xlsx")
            self.end_headers()
            self.wfile.write(data)
            return
        return self.json({"error": "Ruta no encontrada"}, 404)

    def do_POST(self):
        parsed = urlparse(self.path)
        length = int(self.headers.get("Content-Length", "0"))
        ctype = self.headers.get("Content-Type", "")
        if parsed.path == "/api/login":
            raw = self.rfile.read(length) if length else b""
            form = parse_qs(raw.decode("utf-8"))
            user = (form.get("user", [""])[0] or "").strip()
            password = form.get("password", [""])[0] or ""
            with conn() as db:
                user_row = db.execute("SELECT * FROM users WHERE username=? AND active=1", (user,)).fetchone()
            if user_row and verify_password(password, user_row["password_hash"]):
                return self._send(
                    303,
                    "text/plain; charset=utf-8",
                    "OK",
                    {
                        "Set-Cookie": f"crm_session={user_row['username']}|{session_token(user_row['username'], user_row['password_hash'])}; Path=/; HttpOnly; SameSite=Lax",
                        "Location": "/",
                    },
                )
            return self._send(401, "text/html; charset=utf-8", login_page("Usuario o contraseña incorrectos"))
        if parsed.path == "/api/logout":
            return self._send(
                303,
                "text/plain; charset=utf-8",
                "OK",
                {
                    "Set-Cookie": "crm_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax",
                    "Location": "/login",
                },
            )
        if not self.require_auth(parsed.path):
            return
        if parsed.path == "/api/user":
            raw = self.rfile.read(length) if length else b"{}"
            try:
                data = json.loads(raw.decode("utf-8") or "{}")
            except Exception:
                return self.json({"error": "JSON invÃ¡lido"}, 400)
            current = self.current_user()
            if not current or current["role"] != "Administrador":
                return self.json({"error": "Solo administrador"}, 403)
            username = (data.get("username") or "").strip()
            full_name = (data.get("full_name") or username).strip()
            role = data.get("role") or "Vendedor"
            active = 1 if str(data.get("active", "1")) in ("1", "true", "True", "Activo", "on") else 0
            password = data.get("password") or ""
            if not username:
                return self.json({"error": "Falta usuario"}, 400)
            with conn() as db:
                if data.get("id"):
                    existing = db.execute("SELECT * FROM users WHERE id=?", (data.get("id"),)).fetchone()
                    if not existing:
                        return self.json({"error": "Usuario no encontrado"}, 404)
                    if password:
                        db.execute(
                            "UPDATE users SET username=?, full_name=?, role=?, active=?, password_hash=?, updated_at=? WHERE id=?",
                            (username, full_name, role, active, hash_password(password), now_iso(), data.get("id")),
                        )
                    else:
                        db.execute(
                            "UPDATE users SET username=?, full_name=?, role=?, active=?, updated_at=? WHERE id=?",
                            (username, full_name, role, active, now_iso(), data.get("id")),
                        )
                    uid = data.get("id")
                else:
                    if not password:
                        return self.json({"error": "Falta contraseÃ±a"}, 400)
                    cur = db.execute(
                        """
                        INSERT INTO users (username, password_hash, full_name, role, active, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (username, hash_password(password), full_name, role, active, now_iso(), now_iso()),
                    )
                    uid = cur.lastrowid
                db.commit()
            return self.json({"ok": True, "id": uid})
        if parsed.path == "/api/product":
            raw = self.rfile.read(length) if length else b"{}"
            try:
                data = json.loads(raw.decode("utf-8") or "{}")
            except Exception:
                return self.json({"error": "JSON invalido"}, 400)
            current = self.current_user()
            if not is_admin(current):
                return self.json({"error": "Solo el administrador puede modificar productos"}, 403)
            nombre = (data.get("nombre") or "").strip()
            if not nombre:
                return self.json({"error": "Falta el nombre del producto"}, 400)
            precio = data.get("precio")
            precio = float(precio) if precio not in (None, "") else None
            activo = 1 if str(data.get("activo", "1")) in ("1", "true", "True", "on") else 0
            with conn() as db:
                if data.get("id"):
                    db.execute(
                        "UPDATE products SET nombre=?,medidas=?,resistencia=?,peso_aprox=?,precio=?,activo=?,updated_at=? WHERE id=?",
                        (nombre, data.get("medidas", ""), data.get("resistencia", ""), data.get("peso_aprox", ""), precio, activo, now_iso(), data["id"]),
                    )
                else:
                    db.execute(
                        "INSERT INTO products(nombre,medidas,resistencia,peso_aprox,precio,activo,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?)",
                        (nombre, data.get("medidas", ""), data.get("resistencia", ""), data.get("peso_aprox", ""), precio, activo, now_iso(), now_iso()),
                    )
                db.execute(
                    "INSERT INTO product_prices(producto,precio,updated_at) VALUES(?,?,?) ON CONFLICT(producto) DO UPDATE SET precio=excluded.precio,updated_at=excluded.updated_at",
                    (nombre, precio, now_iso()),
                )
                db.commit()
            return self.json({"ok": True})
        if parsed.path == "/api/import_excel":
            if not is_admin(self.current_user()):
                return self.json({"error": "Solo el administrador puede importar archivos"}, 403)
            if cgi is not None:
                env = {"REQUEST_METHOD": "POST", "CONTENT_TYPE": ctype}
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ=env)
                fileitem = form["file"] if "file" in form else None
                if not fileitem:
                    return self.json({"error": "No se recibió archivo"}, 400)
                data = fileitem.file.read()
                filename = fileitem.filename or "Carga manual"
            else:
                raw_upload = self.rfile.read(length) if length else b""
                data, filename = parse_upload_file(raw_upload, ctype)
                if not data:
                    return self.json({"error": "No se recibió archivo"}, 400)
            result = import_excel(io.BytesIO(data), filename)
            return self.json(result)
        raw = self.rfile.read(length) if length else b"{}"
        try:
            data = json.loads(raw.decode("utf-8") or "{}")
        except Exception:
            return self.json({"error": "JSON inválido"}, 400)
        actor = public_user(self.current_user()) or {}
        actor_username = actor.get("username") or ""
        actor_name = actor.get("full_name") or actor_username

        def stamp_actor(record):
            record["created_by"] = actor_username
            record["created_by_name"] = actor_name
            return record

        if parsed.path == "/api/shutdown":
            return self.json({"error": "Apagado deshabilitado en servidor. Usa Cerrar sesion."}, 403)
        if parsed.path == "/api/client":
            stamp_actor(data)
            with conn() as db:
                if data.get("id") and not can_access_client(db, self.current_user(), data.get("id")):
                    return self.json({"error": "Cliente no encontrado o sin permiso"}, 404)
                if not is_admin(self.current_user()):
                    data["assigned_to"] = actor_username
                    if not data.get("id"):
                        scope, params = owner_sql(self.current_user())
                        phone = (data.get("telefono") or data.get("whatsapp") or "").strip()
                        name = (data.get("nombre_comercial") or "").strip()
                        own = db.execute(
                            f"SELECT id FROM clients WHERE ({scope}) AND (lower(nombre_comercial)=lower(?) OR (?<>'' AND (telefono=? OR whatsapp=?))) LIMIT 1",
                            params + [name, phone, phone, phone],
                        ).fetchone()
                        if own:
                            data["id"] = own["id"]
                        else:
                            data["_force_new"] = True
                cid = upsert_client(db, data)
                db.commit()
            return self.json({"ok": True, "id": cid})
        if parsed.path == "/api/interaction":
            if data.get("id"):
                with conn() as db:
                    old = db.execute("SELECT client_id FROM interactions WHERE id=?", (data.get("id"),)).fetchone()
                    if not old or not can_access_client(db, self.current_user(), old["client_id"]):
                        return self.json({"error": "Historial no encontrado o sin permiso"}, 404)
                    cid = update_interaction(db, data.get("id"), data)
                    if not cid:
                        return self.json({"error": "Historial no encontrado"}, 404)
                    db.commit()
                return self.json({"ok": True, "client_id": cid})
            cid = data.get("client_id")
            if not cid:
                return self.json({"error": "Falta cliente"}, 400)
            stamp_actor(data)
            with conn() as db:
                if not can_access_client(db, self.current_user(), cid):
                    return self.json({"error": "Cliente no encontrado o sin permiso"}, 404)
                add_interaction(db, cid, data)
                if data.get("etapa"):
                    db.execute("UPDATE clients SET etapa=?, updated_at=? WHERE id=?", (data["etapa"], now_iso(), cid))
                db.commit()
            return self.json({"ok": True})
        if parsed.path == "/api/quote":
            cid = data.get("client_id")
            if not cid:
                return self.json({"error": "Falta cliente"}, 400)
            cantidad = float(data.get("cantidad") or 0)
            precio = float(data.get("precio_unitario") or 0)
            km_flete = float(data.get("km_flete") or 0)
            tarifa_km = float(data.get("tarifa_km") or 0)
            casetas = float(data.get("casetas") or 0)
            flete_base = km_flete * tarifa_km
            flete_total = flete_base + casetas
            flete_unitario = flete_total / cantidad if cantidad else 0
            costo_individual = precio + flete_unitario
            total = (precio * cantidad) + flete_total
            requiere_factura = str(data.get("requiere_factura", "")).lower() in ("1", "si", "sí", "true", "on")
            subtotal = round(total / 1.16, 2) if requiere_factura else total
            iva = round(total - subtotal, 2) if requiere_factura else 0
            total_con_iva = total
            data["flete_base"] = flete_base
            data["flete_total"] = flete_total
            data["flete_unitario"] = flete_unitario
            data["casetas"] = casetas
            data["precio_ofrecido"] = data.get("precio_ofrecido") or costo_individual
            data["total"] = total
            data["requiere_factura"] = 1 if requiere_factura else 0
            data["subtotal"] = subtotal
            data["iva"] = iva
            data["total_con_iva"] = total_con_iva
            data["tipo_contacto"] = "Cotización"
            data["resultado"] = data.get("resultado") or "Cotización enviada"
            data["etapa"] = data.get("etapa") or "Cotización enviada"
            stamp_actor(data)
            with conn() as db:
                if not can_access_client(db, self.current_user(), cid):
                    return self.json({"error": "Cliente no encontrado o sin permiso"}, 404)
                client = db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
                if not client:
                    return self.json({"error": "Cliente no encontrado"}, 404)
                data["folio"] = build_quote_folio(db, data.get("fecha") or date.today().isoformat(), data.get("producto") or data.get("productos_ofrecidos") or "")
                add_interaction(db, cid, data)
                db.execute("UPDATE clients SET etapa=?, updated_at=? WHERE id=?", ("Cotización enviada", now_iso(), cid))
                db.commit()
            filename = generate_quote_pdf(client, data)
            pdf_url = f"/cotizaciones_pdf/{filename}" if filename else ""
            email_result = {"sent": False}
            if str(data.get("enviar_correo", "")).lower() in ("1", "si", "sÃ­", "true", "on"):
                email_result = send_quote_email(client, data, QUOTE_DIR / filename) if filename else {"sent": False, "error": "No se genero PDF"}
            with conn() as db:
                db.execute(
                    """
                    INSERT INTO quotes
                    (client_id, folio, fecha, cliente, producto, cantidad, precio_unitario, km_flete, tarifa_km, flete_base, flete_total, flete_unitario, casetas, total, requiere_factura, subtotal, iva, total_con_iva, costo_individual, status, pdf, created_at, updated_at, created_by, created_by_name)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        cid,
                        data.get("folio") or "",
                        data.get("fecha") or date.today().isoformat(),
                        client["nombre_comercial"],
                        data.get("producto") or data.get("productos_ofrecidos") or "",
                        cantidad,
                        precio,
                        km_flete,
                        tarifa_km,
                        flete_base,
                        flete_total,
                        flete_unitario,
                        casetas,
                        total,
                        1 if requiere_factura else 0,
                        subtotal,
                        iva,
                        total_con_iva,
                        costo_individual,
                        "Elaborada",
                        pdf_url,
                        now_iso(),
                        now_iso(),
                        actor_username,
                        actor_name,
                    ),
                )
                db.commit()
            return self.json({"ok": True, "pdf": pdf_url, "folio": data.get("folio") or "", "email": email_result, "whatsapp": ""})

        if parsed.path == "/api/quote_status":
            qid = data.get("id")
            status = data.get("status") or "Elaborada"
            if not qid:
                return self.json({"error": "Falta cotizacion"}, 400)
            with conn() as db:
                if not can_access_quote(db, self.current_user(), qid):
                    return self.json({"error": "Cotizacion no encontrada o sin permiso"}, 404)
                quote = db.execute("SELECT * FROM quotes WHERE id=?", (qid,)).fetchone()
                if not quote:
                    return self.json({"error": "Cotizacion no encontrada"}, 404)
                pdf_url = quote["pdf"] or ""
                if status == "Lograda" and pdf_url.startswith("/cotizaciones_pdf/"):
                    source = QUOTE_DIR / Path(pdf_url).name
                    target = QUOTE_WON_DIR / source.name
                    if source.exists():
                        shutil.copy2(source, target)
                        pdf_url = f"/cotizaciones_logradas/{target.name}"
                    db.execute("UPDATE clients SET etapa=?, updated_at=? WHERE id=?", ("Cliente activo", now_iso(), quote["client_id"]))
                db.execute("UPDATE quotes SET status=?, pdf=?, updated_at=? WHERE id=?", (status, pdf_url, now_iso(), qid))
                db.commit()
            return self.json({"ok": True, "status": status, "pdf": pdf_url})

        if parsed.path == "/api/settings":
            if not is_admin(self.current_user()):
                return self.json({"error": "Solo el administrador puede cambiar la configuracion"}, 403)
            with conn() as db:
                for k, v in (data.get("settings") or {}).items():
                    db.execute("INSERT INTO app_settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (k, str(v)))
                for prod, precio in (data.get("productos") or {}).items():
                    db.execute("INSERT INTO product_prices(producto,precio,updated_at) VALUES(?,?,?) ON CONFLICT(producto) DO UPDATE SET precio=excluded.precio, updated_at=excluded.updated_at", (prod, precio if precio != '' else None, now_iso()))
                    db.execute("UPDATE products SET precio=?, updated_at=? WHERE nombre=?", (precio if precio != '' else None, now_iso(), prod))
                for zona, tarifa in (data.get("tarifas") or {}).items():
                    db.execute("INSERT INTO freight_rates(zona,tarifa,updated_at) VALUES(?,?,?) ON CONFLICT(zona) DO UPDATE SET tarifa=excluded.tarifa, updated_at=excluded.updated_at", (zona, tarifa if tarifa != '' else None, now_iso()))
                db.commit()
            return self.json({"ok": True})
        if parsed.path == "/api/monthly_sales":
            if not is_admin(self.current_user()):
                return self.json({"error": "Solo el administrador puede cambiar proyecciones"}, 403)
            year = int(data.get("year") or date.today().year)
            with conn() as db:
                for row in data.get("monthlySales", []):
                    month = int(row.get("month") or 0)
                    if 1 <= month <= 12:
                        amount = float(row.get("amount") or 0)
                        goal = float(row.get("goal") or 0)
                        db.execute("INSERT INTO monthly_sales(year,month,amount,goal,updated_at) VALUES(?,?,?,?,?) ON CONFLICT(year,month) DO UPDATE SET amount=excluded.amount, goal=excluded.goal, updated_at=excluded.updated_at", (year, month, amount, goal, now_iso()))
                db.commit()
            return self.json({"ok": True})
        if parsed.path == "/api/quick":
            stamp_actor(data)
            with conn() as db:
                if not is_admin(self.current_user()):
                    data["assigned_to"] = actor_username
                    data["_force_new"] = True
                cid = upsert_client(db, data)
                add_interaction(db, cid, data)
                db.commit()
            return self.json({"ok": True, "id": cid})
        return self.json({"error": "Ruta no encontrada"}, 404)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        if not self.require_auth(parsed.path):
            return
        if parsed.path == "/api/user":
            current = self.current_user()
            if not current or current["role"] != "Administrador":
                return self.json({"error": "Solo administrador"}, 403)
            uid = qs.get("id", [""])[0]
            if not uid:
                return self.json({"error": "Falta ID de usuario"}, 400)
            with conn() as db:
                row = db.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
                if not row:
                    return self.json({"error": "Usuario no encontrado"}, 404)
                if row["username"] == current["username"]:
                    return self.json({"error": "No puedes desactivar tu propio usuario"}, 400)
                db.execute("UPDATE users SET active=0, updated_at=? WHERE id=?", (now_iso(), uid))
                db.commit()
            return self.json({"ok": True})
        if parsed.path == "/api/client":
            cid = qs.get("id", [""])[0]
            if not cid:
                return self.json({"error": "Falta ID de cliente"}, 400)
            with conn() as db:
                if not can_access_client(db, self.current_user(), cid):
                    return self.json({"error": "Cliente no encontrado o sin permiso"}, 404)
                db.execute("DELETE FROM interactions WHERE client_id=?", (cid,))
                cur = db.execute("DELETE FROM clients WHERE id=?", (cid,))
                db.commit()
            return self.json({"ok": True, "deleted": cur.rowcount})
        if parsed.path == "/api/interaction":
            iid = qs.get("id", [""])[0]
            if not iid:
                return self.json({"error": "Falta ID de historial"}, 400)
            with conn() as db:
                row = db.execute("SELECT client_id FROM interactions WHERE id=?", (iid,)).fetchone()
                if not row or not can_access_client(db, self.current_user(), row["client_id"]):
                    return self.json({"error": "Historial no encontrado o sin permiso"}, 404)
                cur = db.execute("DELETE FROM interactions WHERE id=?", (iid,))
                db.execute("UPDATE clients SET updated_at=? WHERE id=?", (now_iso(), row["client_id"]))
                db.commit()
            return self.json({"ok": True, "deleted": cur.rowcount, "client_id": row["client_id"]})
        return self.json({"error": "Ruta no encontrada"}, 404)


def main():
    bootstrap_persistent_data()
    init_db()
    seed_initial_data()
    auto_backup()
    host = "0.0.0.0"
    port = int(os.environ.get("PORT", "8765"))
    print(f"CRM Bazar de Prefabricados listo en http://127.0.0.1:{port}")
    print("Para celular: abre la IP de esta computadora en la misma red, puerto 8765.")
    ThreadingHTTPServer((host, port), Handler).serve_forever()


if __name__ == "__main__":
    main()
