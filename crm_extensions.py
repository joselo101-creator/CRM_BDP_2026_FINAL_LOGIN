import csv
import io
import json
import shutil
import sqlite3
import threading
import time
import uuid
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except Exception:
    openpyxl = None


EXPORT_TABLES = [
    "clients", "interactions", "quotes", "quote_items", "users",
    "products", "product_prices", "freight_rates", "whatsapp_reminders",
]
SYNC_TABLES = ["clients", "interactions", "quotes", "quote_items", "whatsapp_reminders"]
SYNC_ORDER = ["clients", "interactions", "quotes", "quote_items", "whatsapp_reminders"]


def iso_now():
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def rows_as_dicts(rows):
    return [dict(row) for row in rows]


def safe_name(value):
    return "".join(ch for ch in str(value or "") if ch.isalnum() or ch in ("-", "_", "."))


class CRMExtensions:
    def __init__(self, db_path, data_dir, backup_dir, quote_dir, quote_won_dir):
        self.db_path = Path(db_path)
        self.data_dir = Path(data_dir)
        self.backup_dir = Path(backup_dir)
        self.quote_dir = Path(quote_dir)
        self.quote_won_dir = Path(quote_won_dir)
        self._scheduler_started = False

    def connect(self):
        db = sqlite3.connect(self.db_path, timeout=30)
        db.row_factory = sqlite3.Row
        return db

    @staticmethod
    def _columns(db, table):
        return {row[1] for row in db.execute(f'PRAGMA table_info("{table}")').fetchall()}

    @staticmethod
    def _has_table(db, table):
        return db.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone() is not None

    def ensure_schema(self, db):
        db.executescript("""
            CREATE TABLE IF NOT EXISTS backup_history (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              filename TEXT NOT NULL,
              backup_type TEXT DEFAULT 'Manual',
              created_at TEXT NOT NULL,
              created_by TEXT DEFAULT '',
              size_bytes INTEGER DEFAULT 0,
              notes TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS whatsapp_reminders (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              sync_uuid TEXT UNIQUE,
              fecha_sugerida TEXT DEFAULT '',
              dia TEXT DEFAULT '',
              categoria TEXT DEFAULT '',
              tipo TEXT DEFAULT '',
              objetivo TEXT DEFAULT '',
              mensaje TEXT NOT NULL,
              call_to_action TEXT DEFAULT '',
              estatus TEXT DEFAULT 'Pendiente',
              canal TEXT DEFAULT 'Estado WhatsApp',
              marca TEXT DEFAULT 'Bazar de Prefabricados',
              etiquetas TEXT DEFAULT '',
              client_id INTEGER,
              assigned_to TEXT DEFAULT '',
              media_reference TEXT DEFAULT '',
              fecha_publicacion TEXT DEFAULT '',
              created_at TEXT DEFAULT '',
              updated_at TEXT DEFAULT '',
              created_by TEXT DEFAULT '',
              created_by_name TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS sync_conflicts (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              table_name TEXT NOT NULL,
              sync_uuid TEXT DEFAULT '',
              detected_at TEXT NOT NULL,
              imported_by TEXT DEFAULT '',
              local_json TEXT DEFAULT '',
              incoming_json TEXT DEFAULT '',
              status TEXT DEFAULT 'Pendiente'
            );
        """)
        for table in SYNC_TABLES:
            if not self._has_table(db, table):
                continue
            cols = self._columns(db, table)
            if "sync_uuid" not in cols:
                db.execute(f'ALTER TABLE "{table}" ADD COLUMN sync_uuid TEXT DEFAULT ""')
            if "updated_at" not in cols:
                db.execute(f'ALTER TABLE "{table}" ADD COLUMN updated_at TEXT DEFAULT ""')
            rows = db.execute(f'SELECT id, sync_uuid FROM "{table}" WHERE sync_uuid IS NULL OR sync_uuid=""').fetchall()
            for row in rows:
                db.execute(f'UPDATE "{table}" SET sync_uuid=? WHERE id=?', (str(uuid.uuid4()), row["id"]))
            db.execute(f'DROP INDEX IF EXISTS idx_{table}_sync_uuid')
            try:
                db.execute(
                    f'CREATE UNIQUE INDEX IF NOT EXISTS idx_{table}_sync_uuid '
                    f'ON "{table}"(sync_uuid) WHERE sync_uuid IS NOT NULL AND sync_uuid<>""'
                )
            except sqlite3.IntegrityError:
                pass
        db.execute("INSERT OR IGNORE INTO app_settings(key,value) VALUES('instance_mode',?)", ("local" if self.data_dir == self.db_path.parent and self.data_dir.name != "data" else "online",))
        db.execute("INSERT OR IGNORE INTO app_settings(key,value) VALUES('emergency_base_snapshot_at','')")
        db.commit()

    def _sqlite_snapshot(self, target):
        target = Path(target)
        target.parent.mkdir(parents=True, exist_ok=True)
        source_db = sqlite3.connect(self.db_path, timeout=30)
        destination_db = sqlite3.connect(target)
        try:
            source_db.backup(destination_db)
        finally:
            destination_db.close()
            source_db.close()

    def create_backup(self, backup_type="Manual", username="sistema", notes=""):
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"CRM_BDP_RESPALDO_{backup_type.upper()}_{stamp}.zip"
        target = self.backup_dir / filename
        temp_db = self.backup_dir / f".snapshot_{stamp}.sqlite"
        self._sqlite_snapshot(temp_db)
        manifest = {
            "format": "CRM_BDP_BACKUP_V1",
            "created_at": iso_now(),
            "created_by": username,
            "backup_type": backup_type,
            "database": "bdp_crm.sqlite",
        }
        with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.write(temp_db, "bdp_crm.sqlite")
            archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
            for folder, label in ((self.quote_dir, "cotizaciones_pdf_actuales"), (self.quote_won_dir, "cotizaciones_logradas")):
                if folder.exists():
                    for item in folder.iterdir():
                        if item.is_file():
                            archive.write(item, f"{label}/{item.name}")
        temp_db.unlink(missing_ok=True)
        size = target.stat().st_size
        with self.connect() as db:
            self.ensure_schema(db)
            db.execute(
                "INSERT INTO backup_history(filename,backup_type,created_at,created_by,size_bytes,notes) VALUES(?,?,?,?,?,?)",
                (filename, backup_type, iso_now(), username, size, notes),
            )
            db.commit()
        return {"filename": filename, "size_bytes": size, "created_at": iso_now()}

    def list_backups(self):
        with self.connect() as db:
            self.ensure_schema(db)
            return rows_as_dicts(db.execute("SELECT * FROM backup_history ORDER BY created_at DESC, id DESC LIMIT 100").fetchall())

    def backup_file(self, name):
        clean = safe_name(name)
        target = self.backup_dir / clean
        return target if clean and target.exists() and target.is_file() else None

    def restore_backup_bytes(self, payload, username="admin"):
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        incoming_zip = self.backup_dir / f".restore_{stamp}.zip"
        extract_dir = self.backup_dir / f".restore_{stamp}"
        incoming_zip.write_bytes(payload)
        try:
            with zipfile.ZipFile(incoming_zip, "r") as archive:
                names = set(archive.namelist())
                if "bdp_crm.sqlite" not in names:
                    raise ValueError("El ZIP no contiene bdp_crm.sqlite")
                archive.extractall(extract_dir)
            incoming_db = extract_dir / "bdp_crm.sqlite"
            check = sqlite3.connect(incoming_db)
            try:
                required = {"clients", "interactions", "quotes"}
                tables = {row[0] for row in check.execute("SELECT name FROM sqlite_master WHERE type='table'")}
                if not required.issubset(tables):
                    raise ValueError("La base no corresponde al CRM BDP")
                integrity = check.execute("PRAGMA integrity_check").fetchone()[0]
                if integrity != "ok":
                    raise ValueError("La base no paso la revision de integridad")
            finally:
                check.close()
            pre = self.create_backup("PreRestauracion", username, "Generado antes de restaurar")
            source = sqlite3.connect(incoming_db)
            destination = sqlite3.connect(self.db_path)
            try:
                source.backup(destination)
            finally:
                destination.close()
                source.close()
            for folder_name, destination_dir in (("cotizaciones_pdf_actuales", self.quote_dir), ("cotizaciones_logradas", self.quote_won_dir)):
                source_dir = extract_dir / folder_name
                destination_dir.mkdir(parents=True, exist_ok=True)
                if source_dir.exists():
                    for item in source_dir.iterdir():
                        if item.is_file():
                            shutil.copy2(item, destination_dir / item.name)
            with self.connect() as db:
                self.ensure_schema(db)
                db.execute("INSERT OR REPLACE INTO app_settings(key,value) VALUES('emergency_base_snapshot_at',?)", (iso_now(),))
                db.commit()
            return {"ok": True, "pre_restore_backup": pre["filename"]}
        finally:
            incoming_zip.unlink(missing_ok=True)
            shutil.rmtree(extract_dir, ignore_errors=True)

    def _table_rows(self, db, table):
        if not self._has_table(db, table):
            return []
        return rows_as_dicts(db.execute(f'SELECT * FROM "{table}" ORDER BY rowid').fetchall())

    def export_json(self):
        with self.connect() as db:
            self.ensure_schema(db)
            payload = {"format": "CRM_BDP_EXPORT_V1", "exported_at": iso_now(), "tables": {}}
            for table in EXPORT_TABLES:
                payload["tables"][table] = self._table_rows(db, table)
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    def export_csv_zip(self):
        output = io.BytesIO()
        with self.connect() as db, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
            self.ensure_schema(db)
            for table in EXPORT_TABLES:
                rows = self._table_rows(db, table)
                text = io.StringIO(newline="")
                if rows:
                    writer = csv.DictWriter(text, fieldnames=list(rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(rows)
                archive.writestr(f"{table}.csv", text.getvalue().encode("utf-8-sig"))
        return output.getvalue()

    def export_excel(self):
        if openpyxl is None:
            raise RuntimeError("openpyxl no esta instalado")
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        with self.connect() as db:
            self.ensure_schema(db)
            for table in EXPORT_TABLES:
                sheet = workbook.create_sheet(title=table[:31])
                rows = self._table_rows(db, table)
                if not rows:
                    sheet.append(["Sin registros"])
                    continue
                headers = list(rows[0].keys())
                sheet.append(headers)
                for row in rows:
                    sheet.append([row.get(header) for header in headers])
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions
                for cell in sheet[1]:
                    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8F754D")
                for column in sheet.columns:
                    letter = column[0].column_letter
                    width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
                    sheet.column_dimensions[letter].width = max(width, 12)
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def reminder_summary(self):
        today = date.today().isoformat()
        future = (date.today() + timedelta(days=7)).isoformat()
        with self.connect() as db:
            self.ensure_schema(db)
            today_count = db.execute("SELECT COUNT(*) FROM whatsapp_reminders WHERE fecha_sugerida=? AND estatus NOT IN ('Publicado','Descartado')", (today,)).fetchone()[0]
            overdue = db.execute("SELECT COUNT(*) FROM whatsapp_reminders WHERE fecha_sugerida<? AND estatus NOT IN ('Publicado','Descartado')", (today,)).fetchone()[0]
            upcoming = db.execute("SELECT COUNT(*) FROM whatsapp_reminders WHERE fecha_sugerida>? AND fecha_sugerida<=? AND estatus NOT IN ('Publicado','Descartado')", (today, future)).fetchone()[0]
            last = db.execute("SELECT * FROM backup_history ORDER BY created_at DESC, id DESC LIMIT 1").fetchone()
            mode_row = db.execute("SELECT value FROM app_settings WHERE key='instance_mode'").fetchone()
        return {"today": today_count, "overdue": overdue, "upcoming": upcoming, "last_backup": dict(last) if last else None, "mode": (mode_row[0] if mode_row else "online")}

    def list_reminders(self, filters=None):
        filters = filters or {}
        where = ["1=1"]
        params = []
        mapping = {"status": "r.estatus", "date": "r.fecha_sugerida", "seller": "r.assigned_to", "category": "r.categoria", "client": "r.client_id"}
        for key, column in mapping.items():
            value = filters.get(key)
            if value:
                where.append(f"{column}=?")
                params.append(value)
        with self.connect() as db:
            self.ensure_schema(db)
            rows = db.execute(f"""
                SELECT r.*, c.nombre_comercial AS cliente_nombre, u.full_name AS vendedor_nombre
                FROM whatsapp_reminders r
                LEFT JOIN clients c ON c.id=r.client_id
                LEFT JOIN users u ON u.username=r.assigned_to
                WHERE {' AND '.join(where)}
                ORDER BY r.fecha_sugerida, r.id
                LIMIT 1000
            """, params).fetchall()
        return rows_as_dicts(rows)

    def save_reminder(self, data, actor):
        now = iso_now()
        fields = ["fecha_sugerida", "dia", "categoria", "tipo", "objetivo", "mensaje", "call_to_action", "estatus", "canal", "marca", "etiquetas", "client_id", "assigned_to", "media_reference", "fecha_publicacion"]
        values = {field: data.get(field, "") for field in fields}
        if not str(values["mensaje"] or "").strip():
            raise ValueError("El mensaje es obligatorio")
        if values["fecha_sugerida"]:
            try:
                values["dia"] = datetime.strptime(values["fecha_sugerida"], "%Y-%m-%d").strftime("%A")
            except ValueError:
                pass
        with self.connect() as db:
            self.ensure_schema(db)
            if data.get("id"):
                assignments = ",".join(f"{field}=?" for field in fields)
                db.execute(f"UPDATE whatsapp_reminders SET {assignments}, updated_at=? WHERE id=?", [values[field] for field in fields] + [now, data["id"]])
                rid = int(data["id"])
            else:
                cols = ["sync_uuid"] + fields + ["created_at", "updated_at", "created_by", "created_by_name"]
                placeholders = ",".join("?" for _ in cols)
                vals = [str(uuid.uuid4())] + [values[field] for field in fields] + [now, now, actor.get("username", ""), actor.get("full_name", "")]
                cur = db.execute(f"INSERT INTO whatsapp_reminders({','.join(cols)}) VALUES({placeholders})", vals)
                rid = cur.lastrowid
            db.commit()
        return rid

    def delete_reminder(self, reminder_id):
        with self.connect() as db:
            cur = db.execute("DELETE FROM whatsapp_reminders WHERE id=?", (reminder_id,))
            db.commit()
            return cur.rowcount

    def import_reminders_excel(self, payload, actor):
        if openpyxl is None:
            raise RuntimeError("openpyxl no esta instalado")
        workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
        accepted = ["fecha_sugerida", "dia", "categoria", "tipo", "objetivo", "mensaje", "call_to_action", "estatus", "canal", "marca", "etiquetas"]
        if "mensaje" not in headers:
            raise ValueError("El Excel necesita una columna llamada mensaje")
        imported = 0
        next_date = date.today()
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            if not str(row.get("mensaje") or "").strip():
                continue
            raw_date = row.get("fecha_sugerida")
            if isinstance(raw_date, datetime):
                raw_date = raw_date.date().isoformat()
            elif isinstance(raw_date, date):
                raw_date = raw_date.isoformat()
            elif raw_date:
                raw_date = str(raw_date)[:10]
            else:
                while next_date.weekday() == 6:
                    next_date += timedelta(days=1)
                raw_date = next_date.isoformat()
                next_date += timedelta(days=1)
            data = {key: row.get(key, "") or "" for key in accepted}
            data["fecha_sugerida"] = raw_date
            data["estatus"] = data.get("estatus") or "Pendiente"
            data["canal"] = data.get("canal") or "Estado WhatsApp"
            data["marca"] = data.get("marca") or "Bazar de Prefabricados"
            self.save_reminder(data, actor)
            imported += 1
        return imported

    def export_reminders_excel(self, filters=None):
        if openpyxl is None:
            raise RuntimeError("openpyxl no esta instalado")
        rows = self.list_reminders(filters)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Recordatorios WhatsApp"
        headers = ["fecha_sugerida", "dia", "categoria", "tipo", "objetivo", "mensaje", "call_to_action", "estatus", "canal", "marca", "etiquetas", "cliente_nombre", "vendedor_nombre", "media_reference", "fecha_publicacion", "created_at", "updated_at"]
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8F754D")
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_sync_bundle(self, actor="local"):
        with self.connect() as db:
            self.ensure_schema(db)
            base_row = db.execute("SELECT value FROM app_settings WHERE key='emergency_base_snapshot_at'").fetchone()
            payload = {
                "format": "CRM_BDP_SYNC_V1",
                "exported_at": iso_now(),
                "exported_by": actor,
                "base_snapshot_at": base_row[0] if base_row else "",
                "tables": {},
            }
            client_uuid = {row["id"]: row["sync_uuid"] for row in db.execute("SELECT id,sync_uuid FROM clients").fetchall()}
            quote_uuid = {row["id"]: row["sync_uuid"] for row in db.execute("SELECT id,sync_uuid FROM quotes").fetchall()}
            for table in SYNC_TABLES:
                rows = self._table_rows(db, table)
                for row in rows:
                    if table in ("interactions", "quotes", "whatsapp_reminders") and row.get("client_id"):
                        row["client_sync_uuid"] = client_uuid.get(row["client_id"], "")
                    if table == "quote_items" and row.get("quote_id"):
                        row["quote_sync_uuid"] = quote_uuid.get(row["quote_id"], "")
                payload["tables"][table] = rows
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    @staticmethod
    def _comparable(row):
        ignored = {"id", "client_id", "quote_id", "client_sync_uuid", "quote_sync_uuid"}
        return {key: value for key, value in row.items() if key not in ignored}

    def import_sync_bundle(self, payload, actor="admin"):
        data = json.loads(payload.decode("utf-8-sig"))
        if data.get("format") != "CRM_BDP_SYNC_V1":
            raise ValueError("El archivo no es una exportacion de emergencia valida")
        base = data.get("base_snapshot_at") or ""
        counts = {"inserted": 0, "updated": 0, "skipped": 0, "conflicts": 0}
        with self.connect() as db:
            self.ensure_schema(db)
            id_maps = {"clients": {}, "quotes": {}}
            for table in SYNC_ORDER:
                if not self._has_table(db, table):
                    continue
                columns = self._columns(db, table)
                for incoming in data.get("tables", {}).get(table, []):
                    incoming = dict(incoming)
                    sync_id = incoming.get("sync_uuid") or str(uuid.uuid4())
                    incoming["sync_uuid"] = sync_id
                    if table in ("interactions", "quotes", "whatsapp_reminders") and incoming.get("client_sync_uuid"):
                        parent = db.execute("SELECT id FROM clients WHERE sync_uuid=?", (incoming["client_sync_uuid"],)).fetchone()
                        incoming["client_id"] = parent[0] if parent else None
                    if table == "quote_items" and incoming.get("quote_sync_uuid"):
                        parent = db.execute("SELECT id FROM quotes WHERE sync_uuid=?", (incoming["quote_sync_uuid"],)).fetchone()
                        incoming["quote_id"] = parent[0] if parent else None
                    existing = db.execute(f'SELECT * FROM "{table}" WHERE sync_uuid=?', (sync_id,)).fetchone()
                    usable = {key: value for key, value in incoming.items() if key in columns and key != "id"}
                    if existing:
                        local = dict(existing)
                        incoming_changed = (incoming.get("updated_at") or incoming.get("created_at") or "") > base if base else True
                        local_changed = (local.get("updated_at") or local.get("created_at") or "") > base if base else False
                        differs = self._comparable(local) != self._comparable(incoming)
                        if incoming_changed and local_changed and differs:
                            local_json = json.dumps(local, ensure_ascii=False, sort_keys=True)
                            incoming_json = json.dumps(incoming, ensure_ascii=False, sort_keys=True)
                            repeated = db.execute(
                                "SELECT id FROM sync_conflicts WHERE table_name=? AND sync_uuid=? AND local_json=? AND incoming_json=? AND status='Pendiente'",
                                (table, sync_id, local_json, incoming_json),
                            ).fetchone()
                            if repeated:
                                counts["skipped"] += 1
                            else:
                                db.execute(
                                    "INSERT INTO sync_conflicts(table_name,sync_uuid,detected_at,imported_by,local_json,incoming_json,status) VALUES(?,?,?,?,?,?,?)",
                                    (table, sync_id, iso_now(), actor, local_json, incoming_json, "Pendiente"),
                                )
                                counts["conflicts"] += 1
                            continue
                        if differs:
                            assignments = ",".join(f'"{key}"=?' for key in usable if key != "sync_uuid")
                            keys = [key for key in usable if key != "sync_uuid"]
                            if keys:
                                db.execute(f'UPDATE "{table}" SET {assignments} WHERE sync_uuid=?', [usable[key] for key in keys] + [sync_id])
                            counts["updated"] += 1
                        else:
                            counts["skipped"] += 1
                    else:
                        keys = list(usable.keys())
                        db.execute(f'INSERT INTO "{table}" ({",".join(keys)}) VALUES ({",".join("?" for _ in keys)})', [usable[key] for key in keys])
                        counts["inserted"] += 1
                    if table in id_maps:
                        row = db.execute(f'SELECT id FROM "{table}" WHERE sync_uuid=?', (sync_id,)).fetchone()
                        if row:
                            id_maps[table][sync_id] = row[0]
            db.commit()
        return counts

    def list_conflicts(self):
        with self.connect() as db:
            self.ensure_schema(db)
            return rows_as_dicts(db.execute("SELECT id,table_name,sync_uuid,detected_at,imported_by,status FROM sync_conflicts ORDER BY id DESC LIMIT 200").fetchall())

    def _automatic_due(self, kind, now):
        prefix = f"CRM_BDP_RESPALDO_{kind.upper()}_{now.strftime('%Y%m%d')}"
        return not any(self.backup_dir.glob(prefix + "*.zip"))

    def automatic_backup_tick(self):
        now = datetime.now()
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        if self._automatic_due("Diario", now):
            self.create_backup("Diario", "sistema")
        if now.weekday() == 0 and self._automatic_due("Semanal", now):
            self.create_backup("Semanal", "sistema")
        if now.day == 1 and self._automatic_due("Mensual", now):
            self.create_backup("Mensual", "sistema")
        self._prune("DIARIO", 14)
        self._prune("SEMANAL", 12)
        self._prune("MENSUAL", 18)

    def _prune(self, kind, keep):
        files = sorted(self.backup_dir.glob(f"CRM_BDP_RESPALDO_{kind}_*.zip"), key=lambda item: item.stat().st_mtime, reverse=True)
        for item in files[keep:]:
            item.unlink(missing_ok=True)

    def start_scheduler(self):
        if self._scheduler_started:
            return
        self._scheduler_started = True

        def worker():
            while True:
                try:
                    self.automatic_backup_tick()
                except Exception as exc:
                    print(f"Respaldo automatico pendiente: {exc}")
                time.sleep(3600)

        threading.Thread(target=worker, daemon=True, name="crm-backups").start()
